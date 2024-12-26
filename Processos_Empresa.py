''' Autor: João Pedro Pandolfi Tedesco (2022)'''

import tkinter.messagebox
import win32com.client as win32
import datetime
import os
import openpyxl
import xlrd
from PIL import ImageGrab
import tkinter as tk
import xlsxwriter
import pandas as pd
import time
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import shutil
from selenium.webdriver.common.by import By
import yfinance as yf
from zipfile import ZipFile

preto_Empresa = '#%02x%02x%02x' % (32, 35, 40)  # paleta de cores da Empresa
amarelo_Empresa = '#%02x%02x%02x' % (220, 255, 1)
branco_Empresa = '#%02x%02x%02x' % (242, 243, 236)

caminho_Empresa = 'Empresa_pastas'

dat = []
root = tk.Tk()  # Aqui, define a tela que pergunta a data de hoje. Recebe ela e coloca numa lista.
canvas1 = tk.Canvas(root, width=400, height=300, relief='raised', bg=preto_Empresa)
canvas1.pack()
v = tk.StringVar()
label1 = tk.Label(root, text='Processos - Empresa', font=('helvetica', 14), bg=preto_Empresa, fg=amarelo_Empresa)
canvas1.create_window(200, 25, window=label1)
label2 = tk.Label(root, text=' Insira a data para as operações. \nEx: AAAAMMDD (sem as barras)', font=('helvetica', 10),
                  bg=preto_Empresa, fg=amarelo_Empresa)
canvas1.create_window(200, 100, window=label2)
ent0 = tk.Entry(root, textvariable=v)
canvas1.create_window(200, 140, window=ent0)


def pegainput():
    value = ent0.get()
    button1.config(text='Pronto!', bg=amarelo_Empresa, fg=preto_Empresa)
    dat.append(value)


button1 = tk.Button(text='Definir Data', command=pegainput, bg=branco_Empresa, fg=preto_Empresa,
                    font=('helvetica', 9, 'bold'))
canvas1.create_window(200, 180, window=button1)
button2 = tk.Button(text='Prosseguir >>>', command=lambda: root.destroy(), bg=amarelo_Empresa, fg=preto_Empresa,
                    font=('helvetica', 9, 'bold'))
canvas1.create_window(300, 180, window=button2)
root.mainloop()

data = dat[0]  # Pega a data da lista e bota ela em vários formatos pra rodar os programas
dia = f'{data[6]}{data[7]}'
mes = f'{data[4]}{data[5]}'
ano = f'{data[0]}{data[1]}{data[2]}{data[3]}'
hoj = datetime.datetime(int(ano), int(mes), int(dia))
if hoj.weekday() == 0:
    ont = hoj - datetime.timedelta(days=3)
    antont = hoj - datetime.timedelta(days=4)
elif hoj.weekday() == 1:
    ont = hoj - datetime.timedelta(days=1)
    antont = hoj - datetime.timedelta(days=4)
else:
    ont = hoj - datetime.timedelta(days=1)  # Ontem e anteontem são os último e penúltimo dias úteis, respectivamente
    antont = hoj - datetime.timedelta(days=2)  # Por isso depende se é segunda-feira ou não
onte = str(ont)
anto = str(antont)
ontem = f'{onte[0]}{onte[1]}{onte[2]}{onte[3]}{onte[5]}{onte[6]}{onte[8]}{onte[9]}'
anteontem = f'{anto[0]}{anto[1]}{anto[2]}{anto[3]}{anto[5]}{anto[6]}{anto[8]}{anto[9]}'
ontem_bonito = f'{onte[8]}{onte[9]}/{onte[5]}{onte[6]}/{onte[0]}{onte[1]}{onte[2]}{onte[3]}'
ontem_arq = f'{onte[8]}{onte[9]}-{onte[5]}{onte[6]}-{onte[0]}{onte[1]}{onte[2]}{onte[3]}'
ontem_n_tao_bonito = f'{onte[8]}{onte[9]}{onte[5]}{onte[6]}{onte[0]}{onte[1]}{onte[2]}{onte[3]}'
yesterday = pd.to_datetime(f'{onte[5]}{onte[6]}/{onte[8]}{onte[9]}/{onte[0]}{onte[1]}{onte[2]}{onte[3]}')

##   Dicionário feito pro sistema

mudancas = {'RL': 3, 'T': 2, 'D': 1, 'RB': 2, 'A': 1, 'RT': 5, 'F': 2, 'XP_INV': 3, 'adminCTVM': 72, 'MODAL': 10197,
            'BTG PACT': 10198, 'STONEX': 10196, 'STONE': 10196, 'ITAU':114, 'cod1': 387894, 'cod2':454109,
            'cod3': 454168,'cod4': 652091, '1982': 10197, '1130':10196, '114':114, '3':3, '72':72, '85':10198,
            '270702': 387894, '502647': 454168, '685779': 652091, 'EMP0012': 454109, 'EMP0030': 387894, 'XP US': 4,
            '0': 72, 'RC': 4}

# FUNÇÕES:


def muda_data(data):
    return f'{data[6]}{data[7]}/{data[4]}{data[5]}/{data[0]}{data[1]}{data[2]}{data[3]}'


def naonul(x):
    if x == 0:
        return '-'
    else:
        return x


def transf_data(xl_date):  # Função que pega a data em formato excel (que é o número absoluto) e transforma pra DD/MM/AAAA
    datetime_date = xlrd.xldate_as_datetime(xl_date, 0)
    date_object = datetime_date.date()
    data = date_object.isoformat()
    return f'{data[8]}{data[9]}/{data[5]}{data[6]}/{data[0]}{data[1]}{data[2]}{data[3]}'


def prox_5():  # Função que retorna a lista dos próximos 5 dias úteis em formato DD/MM/AAAA
    if hoj.weekday() == 0:
        d1 = hoj + datetime.timedelta(days=1)
        d2 = hoj + datetime.timedelta(days=2)
        d3 = hoj + datetime.timedelta(days=3)
        d4 = hoj + datetime.timedelta(days=4)
        d5 = hoj + datetime.timedelta(days=7)
    elif hoj.weekday() == 1:
        d1 = hoj + datetime.timedelta(days=1)
        d2 = hoj + datetime.timedelta(days=2)
        d3 = hoj + datetime.timedelta(days=3)
        d4 = hoj + datetime.timedelta(days=6)
        d5 = hoj + datetime.timedelta(days=7)
    elif hoj.weekday() == 2:
        d1 = hoj + datetime.timedelta(days=1)
        d2 = hoj + datetime.timedelta(days=2)
        d3 = hoj + datetime.timedelta(days=5)
        d4 = hoj + datetime.timedelta(days=6)
        d5 = hoj + datetime.timedelta(days=7)
    elif hoj.weekday() == 3:
        d1 = hoj + datetime.timedelta(days=1)
        d2 = hoj + datetime.timedelta(days=4)
        d3 = hoj + datetime.timedelta(days=5)
        d4 = hoj + datetime.timedelta(days=6)
        d5 = hoj + datetime.timedelta(days=7)
    else:
        d1 = hoj + datetime.timedelta(days=3)
        d2 = hoj + datetime.timedelta(days=4)
        d3 = hoj + datetime.timedelta(days=5)
        d4 = hoj + datetime.timedelta(days=6)
        d5 = hoj + datetime.timedelta(days=7)
    return [hoj, d1, d2, d3, d4, d5]


def le_linha(o):  # Função que lê uma linha de XML e pega só o valor, sem os identificadores.
    try:
        of = o.split('<', 30)
        reso = of[1]
        off = reso.split('>', 30)
        return off[1]
    except:
        pass


def le_xml(data, fundo):
    troca = {'m1':'Fundo_master_1', 'm2':'fundo_master_2', 'f1': 'Fundo_fic_1', 'f2':'Fundo_fic_2'}
    with open(fr'C:\Users\{caminho_Empresa}\ROTINAS\Balanco\Carteiras XML\{troca[fundo]}\{data}.xml') as k:
        linhas = k.readlines()
        codigos = {'2': 'Auditoria', '8': 'Outras Despesas', '999': 'Outros', '13': 'Taxa CETIP', '12': 'Taxa ANBIMA',
                   '14': 'Taxa CVM', '4': 'Cartório',
                   '31': 'Empréstimo de Ações', '21': 'Ações ou Opções', '16': 'Taxa SELIC', '27': 'Dividendos',
                   '28': 'Juros s/ Cap. Prório', '38': 'Emolumentos',
                   '34': 'Taxa ADM', '39': 'Valor Bovespa', '40': 'Valor Repasse Bovespa', '47': 'Resgate a liquidar',
                   '45': 'Aplicação a Converter',
                   '35': 'Pfee', '36': 'Corretagem Bovespa', '15': 'Taxa Custódia', '22': 'Derivativos',
                   '5': 'Correspondências', '24': 'Termo SELIC', '29': 'Subscrições'}
        offshore = 0
        tesouraria = 0
        qtcotazeragem = 0
        cotazeragem = 0
        zeragem = 0
        acoes = []
        opcoes = []
        tit_publico = []
        provisoes = []
        ct = 0
        for i in linhas:
            if '<titpublico>' in i:
                if linhas[ct + 23] == '\t\t<compromisso>\n':
                    zeragem += float(le_linha(linhas[ct + 17]))
                else:
                    venc = le_linha(linhas[ct + 6])
                    tit_publico.append([f'{venc[6:8]}/{venc[4:6]}/{venc[0:4]}', float(le_linha(linhas[ct+14])),
                                        float(le_linha(linhas[ct+16])), float(le_linha(linhas[ct+18]))])
            elif '<acoes>' in i:
                if le_linha(linhas[ct+13]) in ['C', 'D', 'T']:
                    sinal = 1
                else:
                    sinal = -1
                acoes.append([le_linha(linhas[ct + 3]), round(float(le_linha(linhas[ct + 4])), 2) * sinal,
                                  float(le_linha(linhas[ct + 10])),
                                  float(le_linha(linhas[ct + 7])), round(float(le_linha(linhas[ct + 6])), 2),
                                  round(float(le_linha(linhas[ct + 15])), 2), le_linha(linhas[ct + 13])])
            elif '<opcoesacoes>' in i:
                if le_linha(linhas[ct+9]) in ['C', 'D', 'T']:
                    sinal = 1
                else:
                    sinal = -1
                opcoes.append([le_linha(linhas[ct + 3]), le_linha(linhas[ct + 4]), round(float(le_linha(linhas[ct + 5])), 2) * sinal,
                               float(le_linha(linhas[ct + 11])), float(le_linha(linhas[ct + 6])), round(float(le_linha(linhas[ct + 7])), 2), le_linha(linhas[ct + 9]),
                               muda_data(le_linha(linhas[ct + 8]))])
            elif '<provisao>' in i:
                if le_linha(linhas[ct+2]) == 'C':
                    sinal = 1
                else:
                    sinal = -1
                provisoes.append(
                    [codigos[le_linha(linhas[ct+1])], le_linha(linhas[ct+2]), le_linha(linhas[ct+3]), round(float(le_linha(linhas[ct+4])), 2) * sinal])
            elif '<caixa>' in i:
                if le_linha(linhas[ct+1]) == 'BRBBDC':
                    tesouraria += float(le_linha(linhas[ct+3]))
                else:
                    offshore += float(le_linha(linhas[ct + 3]))
            elif '<cotas>' in i:
                if le_linha(linhas[ct + 1]) == 'BRTPF2CTF005':
                    qtcotazeragem = float(le_linha(linhas[ct + 3]))
                    cotazeragem = float(le_linha(linhas[ct + 5]))
                    zeragem += qtcotazeragem * cotazeragem  #caixa será [tesouraria + zeragem, offshore]
            ct += 1
    if len(opcoes) == 0:
        opcoes = [['-', '-', '-', '-', '-', '-', '-', '-']]
    if len(tit_publico) == 0:
        tit_publico = [['', 0, 0, 0]]
    return [le_linha(linhas[15]), le_linha(linhas[16]), le_linha(linhas[17]),
            float(le_linha(linhas[21])) - float(le_linha(linhas[22])), cotazeragem, qtcotazeragem,
            [tesouraria + zeragem, offshore], acoes, opcoes, tit_publico, provisoes]
    # cota, qtcota, pl, cotas a emitir


class Carteira:
    def __init__(self, data, fundo):
        self.data = data
        self.fundo = fundo
        nomes = {'f1': 'FIC1', 'm1': 'Master1', 'f2': 'FIC2', 'm2': 'Master2'}
        self.nome = nomes[fundo]
        xml = le_xml(data, fundo)
        self.cota = xml[0]
        self.qt_cota = xml[1]
        self.pl = xml[2]
        self.cotas_a_emitir = xml[3]
        self.cotazeragem = xml[4]
        self.qtcotazeragem = xml[5]
        self.caixa = xml[6]
        self.acoes = xml[7]
        self.opcoes = xml[8]
        self.titpublico = xml[9]
        self.provisoes = xml[10]

    def btcs(self):
        try:
            btcs = []
            bbi = open(
                rf'C:\Users\{caminho_Empresa}\ROTINAS\Relatório - BTC\Posições BTC - BIP\{self.data}.csv',
                'r').readlines()
            bbi.pop(0)
            for k in bbi:
                contrato = k.split(';', 1000)
                if contrato[4] == 'D':
                    cont = 'Doador'
                else:
                    cont = 'Tomador'
                if contrato[6] == '685779':
                    fundo = 'Master I'
                else:
                    fundo = 'Master II'
                linha_contrato = [contrato[2], cont, fundo, contrato[3], contrato[8], contrato[12], contrato[1],
                                  contrato[16].capitalize()]
                btcs.append(linha_contrato)
            btcs.pop(0)
            return btcs
        except:
            btcs = []
            bbi = open(
                rf'C:\Users\{caminho_Empresa}\ROTINAS\Relatório - BTC\Posições BTC - BBI\{self.data}.csv',
                'r').readlines()
            for k in bbi:
                contrato = k.split(';', 1000)
                if contrato[2] == 'D':
                    cont = 'Doador'
                else:
                    cont = 'Tomador'
                if contrato[1] == '685779':
                    fundo = 'Master I'
                else:
                    fundo = 'Master II'
                linha_contrato = [contrato[4], cont, fundo, contrato[18], contrato[20], contrato[9], contrato[3],
                                  contrato[23]]
                btcs.append(linha_contrato)
            btcs.pop(0)
            return btcs

    def garantias(self):
        try:
            garantias = []
            bbi = open(
                rf'C:\Users\{caminho_Empresa}\ROTINAS\Relatório - BTC\Garantias - BIP\{self.data}.csv',
                'r').readlines()
            for k in bbi:
                contrato = k.split(';', 1000)
                if contrato[2] == '685779':
                    fundo = 'Master I'
                else:
                    fundo = 'Master II'
                try:
                    depositado = float(str(contrato[9]).replace(',', '.'))
                    requerido = depositado - float(str(contrato[16]).replace(',', '.'))
                except:
                    depositado = ''  # Aqui só um tratamento de erro pra quando ele ler a primeira linha, só de letras, não retornar que não conseguiu passá-las pra float
                    requerido = ''
                linha_contrato = [contrato[4], fundo, depositado, requerido]
                garantias.append(linha_contrato)
            garantias.pop(0)
            return garantias
        except:
            garantias = []
            bbi = open(
                rf'C:\Users\{caminho_Empresa}\ROTINAS\Relatório - BTC\Garantias - BBI\{self.data}.csv',
                'r').readlines()
            for k in bbi:
                contrato = k.split(';', 1000)
                if contrato[1] == '685779':
                    fundo = 'Master I'
                else:
                    fundo = 'Master II'
                try:
                    depositado = float(str(contrato[14]).replace(',', '.'))
                    requerido = depositado - float(str(contrato[11]).replace(',', '.'))
                except:
                    depositado = ''  # Aqui só um tratamento de erro pra quando ele ler a primeira linha, só de letras, não retornar que não conseguiu passá-las pra float
                    requerido = ''
                linha_contrato = [contrato[7], fundo, depositado, requerido]
                garantias.append(linha_contrato)
            garantias.pop(0)
            return garantias

    def pega_carteira(self, cambio):
        troca = {'m1': 'Master1', 'm2': 'Master2', 'f1': 'FIC1',
                 'f2': 'FIC2'}
        with open(
                fr'C:\Users\{caminho_Empresa}\ROTINAS\Balanco\Carteiras XML\{troca[self.fundo.replace("m", "f")]}\{self.data}.xml') as k:
            pl_fic = le_linha(k.readlines()[17])
        carteira = pd.DataFrame(self.acoes).rename(
            columns={0: 'Ticker', 1: 'Quantidade', 2: 'Preco', 3: 'Financeiro', 4: 'Bloq', 5: 'TX Aluguel',
                     6: 'Operação'})  # columns=['Ticker', 'Quantidade', 'Preco', 'Financeiro', 'Garantia', 'TX Aluguel']
        dia = str(self.data)
        comprado = carteira[carteira['Operação'] == 'C'].drop(columns=['Operação'])
        vendido = carteira[carteira['Operação'] == 'V'].drop(columns=['Operação'])
        btc_doador = carteira[carteira['Operação'] == 'D'].rename(columns={'Bloq': 'Doado'}).drop(
            columns=['Quantidade', 'Preco', 'Financeiro', 'TX Aluguel', 'Operação']).groupby('Ticker',
                                                                                             as_index=False).sum()
        btc_tomador = carteira[carteira['Operação'] == 'T'].rename(columns={'Bloq': 'Tomado'}).drop(
            columns=['Preco', 'Financeiro', 'Quantidade', 'TX Aluguel', 'Operação']).groupby('Ticker',
                                                                                             as_index=False).sum()
        btc_tomador['Tomado'] = btc_tomador['Tomado'] * -1
        ativos_emp = pd.concat([comprado, vendido])
        ativos_emp = ativos_emp.merge(btc_tomador, how='left').fillna(0)
        ativos_emp = ativos_emp.merge(btc_doador, how='left').fillna(0)
        ativos_emp.insert(0, 'Data', f'{dia[6:8]}/{dia[4:6]}/{dia[0:4]}')
        ativos_emp.insert(7, 'PL', float(pl_fic))
        ativos_emp['Cot dolar'] = ativos_emp['Preco'] / cambio
        ativos_emp['Aluguel'] = ativos_emp['Tomado'] + ativos_emp['Doado']
        ativos_emp['Quantidade'] = ativos_emp['Quantidade'] + ativos_emp['Bloq'] + ativos_emp['Aluguel']
        ativos_emp['Financeiro'] = ativos_emp['Quantidade'] * ativos_emp['Preco']
        ativos_emp['% do PL'] = ativos_emp['Financeiro'] / ativos_emp['PL']
        ativos_emp['% Bloq'] = ativos_emp['Bloq'] / ativos_emp['Quantidade']
        ativos_emp['% Aluguel'] = ativos_emp['Aluguel'] / ativos_emp['Quantidade']
        ativos_emp = ativos_emp.sort_values(by='% do PL', ascending=False)
        return ativos_emp

# bloco que define as variáveis Carteira, que serão usadas nos processos. Caso os arquivos XML estejam na pasta, ele irá alocá-los, do contrário irá avisar que não há XML.
try:  # Isso ainda não está 100%. Pensar em melhorias.
    cart_m1_ont = Carteira(ontem, 'm1')
    cart_m1_antont = Carteira(anteontem, 'm1')
    cart_m2_ont = Carteira(ontem, 'm2')
    cart_m2_antont = Carteira(anteontem, 'm2')
    cart_f1_ont = Carteira(ontem, 'f1')
    cart_f1_antont = Carteira(anteontem, 'f1')
    cart_f2_ont = Carteira(ontem, 'f2')
    cart_f2_antont = Carteira(anteontem, 'f2')
except:
    try:
        caminhos = [os.path.join(
            rf'C:\Users\{caminho_Empresa}\ROTINAS\PASTA INPUT',
            nome) for nome in os.listdir(
            rf'C:\Users\{caminho_Empresa}\ROTINAS\PASTA INPUT')]
        movimentos = [arq for arq in caminhos if os.path.isfile(arq)]
        for i in movimentos:
            if ontem in i and 'Nome_master1' in i and '.XML' in i:
                os.rename(i,
                          fr'C:\Users\{caminho_Empresa}\Master1\{ontem}.xml')
                shutil.copyfile(
                    fr'C:\Users\{caminho_Empresa}\FIC1\{ontem}.xml',
                    fr'C:\Users\{caminho_Empresa}\{ontem[0:4]}.{ontem[4:6]}.{ontem[6:8]}_Master_FIA.xml')
            elif ontem in i and 'Nome_FIC1' in i and '.XML' in i:
                os.rename(i,
                          fr'C:\Users\{caminho_Empresa}\FIC1\{ontem}.xml')
                shutil.copyfile(
                    fr'C:\Users\{caminho_Empresa}\FIC1\{ontem}.xml',
                    fr'C:\Users\{caminho_Empresa}FIC1\{ontem[0:4]}.{ontem[4:6]}.{ontem[6:8]}.xml')
            elif ontem in i and 'Nome_FIC2' in i and '.XML' in i:
                os.rename(i,
                          fr'C:\Users\{caminho_Empresa}\FIC2{ontem}.xml')
                shutil.copyfile(
                    fr'C:\Users\{caminho_Empresa}\FIC2\{ontem}.xml',
                    fr'C:\Users\{caminho_Empresa}\FIC2\{ontem[0:4]}.{ontem[4:6]}.{ontem[6:8]} FIC FIA.xml')
            elif ontem in i and 'Nome_master2' in i and '.XML' in i:
                os.rename(i,
                          fr'C:\Users\{caminho_Empresa}\Master2\{ontem}.xml')
                shutil.copyfile(
                    fr'C:\Users\{caminho_Empresa}\Master2\{ontem}.xml',
                    fr'C:\Users\{caminho_Empresa}\Master2\{ontem[0:4]}.{ontem[4:6]}.{ontem[6:8]} MASTER FIA.xml')
        cart_m1_ont = Carteira(ontem, 'm1')
        cart_m1_antont = Carteira(anteontem, 'm1')
        cart_m2_ont = Carteira(ontem, 'm2')
        cart_m2_antont = Carteira(anteontem, 'm2')
        cart_f1_ont = Carteira(ontem, 'f1')
        cart_f1_antont = Carteira(anteontem, 'f1')
        cart_f2_ont = Carteira(ontem, 'f2')
        cart_f2_antont = Carteira(anteontem, 'f2')
    except:
        tk.messagebox.showinfo('XML diário', "Não há nenhum arquivo XML na pasta referente a data de d-1."
                                         "\n\n Coloque-os na Pasta Input e rode o programa novamente.")


def comando1():
    try:
        movs = pd.read_excel(
            fr'C:\Users\{caminho_Empresa}\ROTINAS\Balanco\Movimentos\Posições D-30\{data}.xls',
            0)
        movs = movs.drop(columns=[a for a in list(movs) if a not in ['CD_FUNDO', 'DS_DT_LIQUIDACAO_FISICA', 'VL_LIQUIDO']])
        movs = movs.groupby(['CD_FUNDO', 'DS_DT_LIQUIDACAO_FISICA'], as_index=False).sum().sort_values(
            by='DS_DT_LIQUIDACAO_FISICA')
        resgs = pd.pivot_table(movs[movs['DS_DT_LIQUIDACAO_FISICA'] >= pd.to_datetime(data)], values='VL_LIQUIDO',
                               columns='CD_FUNDO', index='DS_DT_LIQUIDACAO_FISICA')
        d5 = pd.DataFrame(index=prox_5()).join(resgs).fillna('-').reset_index()
        d5['index'] = [muda_data(str(a).replace('-', '')) for a in list(d5['index'])]
        resgs = resgs[resgs.index > prox_5()[5]].reset_index().fillna('-')
        resgs['DS_DT_LIQUIDACAO_FISICA'] = [muda_data(str(a).replace('-', '')) for a in list(resgs['DS_DT_LIQUIDACAO_FISICA'])]
        car_cap1 = pd.DataFrame(cart_m1_ont.provisoes)
        car_cap1 = car_cap1[car_cap1[0] != 'Empréstimo de Ações'].groupby(2, as_index=False).sum().rename(
            columns={3: 'Valor I'})
        car_cap1[2] = pd.to_datetime(car_cap1[2])
        car_cap2 = pd.DataFrame(cart_m2_ont.provisoes)
        car_cap2 = car_cap2[car_cap2[0] != 'Empréstimo de Ações'].groupby(2, as_index=False).sum().rename(
            columns={3: 'Valor II'})
        car_cap2[2] = pd.to_datetime(car_cap2[2])
        desp5 = pd.DataFrame(index=prox_5()).join([car_cap1.set_index(2), car_cap2.set_index(2)]).fillna(
            '-').reset_index()
        desp5['index'] = [muda_data(str(a).replace('-', '')) for a in list(desp5['index'])]
        wb = openpyxl.load_workbook(
            rf'C:\Users\{caminho_Empresa}\ROTINAS\Balanco\Movimentos\Template - Liquidez.xlsx')
        ws = wb.sheetnames
        planilha = wb[ws[0]]
        s = 1
        for f in list(d5):
            z = 3
            for k in list(d5[f]):
                planilha.cell(row=z, column=s).value = k
                z += 1
            s += 1
        ss = 4
        for ff in list(resgs):
            zz = 3
            for k in list(resgs[ff]):
                planilha.cell(row=zz, column=ss).value = k
                zz += 1
            ss += 1
        sz = 4
        for fc in list(desp5):
            zs = 15
            for k in list(desp5[fc]):
                planilha.cell(row=zs, column=sz).value = k
                zs += 1
            sz += 1
        caixa_m1 = cart_m1_ont.caixa
        caixa_f1 = cart_f1_ont.caixa
        caixa_m2 = cart_m2_ont.caixa
        caixa_f2 = cart_f2_ont.caixa
        planilha.cell(row=12, column=2).value = 'R$  {:,.2f}'.format(caixa_m1[0] + sum([a[1] for a in cart_m1_ont.titpublico]))
        planilha.cell(row=13, column=2).value = 'R$  {:,.2f}'.format(caixa_f1[0] + sum([a[1] for a in cart_f1_ont.titpublico]))
        planilha.cell(row=14, column=2).value = 'R$  {:,.2f}'.format(caixa_m2[0] + sum([a[1] for a in cart_m2_ont.titpublico]))
        planilha.cell(row=15, column=2).value = 'R$  {:,.2f}'.format(caixa_f2[0] + sum([a[1] for a in cart_f1_ont.titpublico]))
        planilha.cell(row=12, column=3).value = 'R$  {:,.2f}'.format(caixa_m1[1])
        planilha.cell(row=14, column=3).value = 'R$  {:,.2f}'.format(caixa_m2[1])
        planilha.cell(row=17, column=2).value = (caixa_m1[0] + sum([a[1] for a in cart_m1_ont.titpublico])) / float(cart_f1_ont.pl)
        planilha.cell(row=18, column=2).value = (caixa_f1[0] + sum([a[1] for a in cart_f1_ont.titpublico])) / float(cart_f1_ont.pl)
        planilha.cell(row=19, column=2).value = (caixa_m2[0] + sum([a[1] for a in cart_m2_ont.titpublico])) / float(cart_f2_ont.pl)
        planilha.cell(row=20, column=2).value = (caixa_f2[0] + sum([a[1] for a in cart_f2_ont.titpublico])) / float(cart_f2_ont.pl)
        planilha.cell(row=17, column=3).value = caixa_m1[1] / float(cart_f1_ont.pl)
        planilha.cell(row=19, column=3).value = caixa_m2[1] / float(cart_f2_ont.pl)
        wb.save(
            rf'C:\Users\{caminho_Empresa}\ROTINAS\Balanco\Movimentos\Enviados\{data} - Email.xlsx')
        excel = win32.Dispatch('Excel.Application')
        wb = excel.Workbooks.Open(
            rf'C:\Users\{caminho_Empresa}\ROTINAS\Balanco\Movimentos\Enviados\{data} - Email.xlsx')
        sheet = wb.Sheets[0]
        excel.visible = 1
        copyrange = sheet.Range('A1:F21')
        copyrange.CopyPicture(Appearance=1, Format=2)
        ImageGrab.grabclipboard().save(
            rf'C:\Users\{caminho_Empresa}\ROTINAS\Balanco\Movimentos\Enviados\{data}-Tabela.png')
        excel.Quit()

        html_body = '''
            <div>
                Prezados, boa tarde! Seguem os relatórios de resgate de hoje:
            </div>
        '''
        outlook = win32.Dispatch('outlook.application')
        mail = outlook.CreateItem(0)
        mail.To = 'mail1'
        mail.CC = 'mail2'
        mail.Subject = f'Resgates Programados - {dia}/{mes}/{ano}'
        mail.HTMLBody = html_body + rf'<br><img src="C:\Users\{caminho_Empresa}\ROTINAS\Balanco\Movimentos\Enviados\{data}-Tabela.png">'
        mail.Display()
    except:
        tk.messagebox.showinfo('Erro', 'Não há nenhum arquivo de movimentações na pasta. Por favor, insira-o.')


def comando2():
    caminhos = [os.path.join(
        rf'C:\Users\{caminho_Empresa}\ROTINAS\Movimentação de Passivo\Arquivos - Recebidos\{data}',
        nome) for nome in os.listdir(
        rf'C:\Users\{caminho_Empresa}\ROTINAS\Movimentação de Passivo\Arquivos - Recebidos\{data}')]
    movimentos = [arq for arq in caminhos if os.path.isfile(arq)]
    lista_mov = []
    lista_btg_ = []
    lista_orama_ = []
    for i in movimentos:
        if 'BTG' in i:
            with open(i) as btg:
                num_aplicacoes = len(btg.readlines()) - 1
                num_rep = 0
                btg.seek(0)
                btg.readline()
                while num_rep < num_aplicacoes:
                    lista_btg_.append(btg.readline())
                    num_rep = num_rep + 1
        elif 'Orama' in i:
            with open(i) as orama:
                num_aplicacoes = len(orama.readlines()) - 1
                num_rep = 0
                orama.seek(0)
                orama.readline()
                while num_rep < num_aplicacoes:
                    lista_orama_.append(orama.readline())
                    num_rep = num_rep + 1
    for j in lista_btg_:
        a = j.split(';', 1000)
        b = a[2]
        if a[4] == "RP":
            op = 'RB'
        else:
            op = a[4]
        if b[22] == "0":
            dig1 = ''
        else:
            dig1 = b[22]
        if b[23] == "0":
            dig2 = ''
        else:
            dig2 = b[23]
        if a[3] == 'cnpj2':
            cod = 'cod2'
        else:
            cod = 'cod1'
        if a[4] == 'RT':
            resg = ''
        else:
            resg = a[6]
        if a[4] == "RP":
            brut = 'I'
        else:
            brut = ''
        lista_mov.append(
            ['', op, 'N', f'empBT-{dig1}{dig2}{b[24]}{b[25]}{b[26]}{b[27]}{b[28]}{b[29]}', cod, f'{dia}/{mes}/{ano}',
             resg, '', '', 'TE', '208', '0001', '426803', brut, 'R', 'STR', 'N', '', '', '', '', '', '', 'N', ''])
    for j in lista_orama_:
        a = j.split(';', 1000)
        b = a[2]
        if a[4] == "RP":
            op = 'RB'
        else:
            op = a[4]
        if b[9] == "0":
            dig1 = ''
        else:
            dig1 = b[9]
        if a[3] == 'cnpj1':
            cod = 'cod2'
        else:
            cod = 'cod1'
        if a[4] == 'RT':
            resg = ''
        else:
            resg = a[6]
        if a[4] == "RP":
            brut = 'I'
        else:
            brut = ''
        lista_mov.append(
            ['', op, 'N', f'EMP-{dig1}{b[10]}{b[11]}{b[12]}{b[13]}{b[14]}{b[15]}{b[16]}', cod, f'{dia}/{mes}/{ano}',
             resg, '', '', 'TE', '325', '0001', '527802', brut, 'R', 'STR', 'N', '', '', '', '', '', '', 'N', ''])
    with open(
            rf'C:\Users\{caminho_Empresa}\ROTINAS\Movimentação de Passivo\Arquivos - Upload Custódia\{data} - Upload Custódia.txt',
            'w') as upload:
        for i in lista_mov:
            for j in i:
                upload.write(f'{j}\t')
            upload.write('\n')

    tk.messagebox.showinfo('Upload - Custódia',
                           'O arquivo se encontra na pasta "Arquivos - Upload", com a data relativa a da operação. \n'
                           'Faça o upload no admin custódia e extraia o arquivo de movimentos')


def comando3():
    dados3 = []
    prompt3 = tk.Tk()
    telamovbon = tk.Canvas(prompt3, width=400, height=300, relief='raised', bg=preto_Empresa)
    telamovbon.pack()
    tit3 = tk.Label(prompt3, text='Movimentação de Passivo - Emails', font=('helvetica', 14), bg=preto_Empresa,
                    fg=amarelo_Empresa)
    telamovbon.create_window(200, 25, window=tit3)
    l31 = tk.Label(prompt3, text='Houve movimentações no dia de hoje?',
                   font=('helvetica', 10), bg=preto_Empresa, fg=amarelo_Empresa)
    telamovbon.create_window(200, 100, window=l31)
    v31 = tk.StringVar(prompt3)
    v31.set('Selecione')
    l32 = tk.OptionMenu(prompt3, v31, *['Sim', 'Não'])
    telamovbon.create_window(200, 140, window=l32)

    def pegainput3():
        global w31
        w31 = v31.get()
        dados3.append(w31)
        peg1.config(text='Pronto!', bg=amarelo_Empresa, fg=preto_Empresa)
        v31.set('Selecione')

    peg1 = tk.Button(prompt3, text='OK', command=pegainput3, bg=branco_Empresa, fg=preto_Empresa,
                     font=('helvetica', 9, 'bold'))
    telamovbon.create_window(200, 180, window=peg1)
    peg2 = tk.Button(prompt3, text='Prosseguir >>>', command=lambda: [f() for f in [prompt3.destroy, root1.quit]],
                     bg=amarelo_Empresa, fg=preto_Empresa,
                     font=('helvetica', 9, 'bold'))
    telamovbon.create_window(300, 180, window=peg2)
    prompt3.mainloop()

    if dados3[0] == "Não":
        outlook = win32.Dispatch('outlook.application')
        mail = outlook.CreateItem(0)
        mail.To = 'mail1'
        mail.CC = 'mail2'
        mail.Subject = f'Movimentação - Fundos - {dia}/{mes}/{ano}'
        mail.Body = 'Boa tarde! \n\nNão houve movimentações no dia de hoje! \n\nAtt,'
        mail.Display()
    else:
        movs_orama = []
        movs_BTG = []
        ops = []
        tabelamovs = xlrd.open_workbook(rf'C:\Users\{caminho_Empresa}\ROTINAS\Movimentação de Passivo\Arquivos - Excel\{data}\Movimentos.xls')
        tabmovs = tabelamovs.sheet_by_index(0)
        for j in range(0, tabmovs.nrows):
            if "ORAMA DTVM S/A " in tabmovs.row_values(j):
                movs_orama.append(tabmovs.row_values(j))
            elif "BANCO BTG PACTUAL S.A " in tabmovs.row_values(j):
                movs_BTG.append(tabmovs.row_values(j))
        for j in range(1, tabmovs.nrows):
            x = tabmovs.row_values(j)
            if x[4] != 'A':
                ops.append([x[4], x[11], x[3], x[1], x[18], transf_data(x[5]), transf_data(x[6]), transf_data(
                    x[7])])  # [operação, valor bruto, fundo resgatado, cotista, qtdcota, datamov, datacot, dataliq]
            else:
                ops.append([x[4], x[11], x[3], x[1], x[13], transf_data(x[5]), transf_data(x[6]), transf_data(
                    x[7])])  # [operação, valor bruto, fundo resgatado, cotista, qtdcota, datamov, datacot, dataliq]
        opstot = pd.DataFrame(ops, columns=['Operação', 'Valor', 'Fundo', 'Cotista', 'Qtdcota', 'Datamov', 'Datacot',
                                            'Dataliq'])
        rt = opstot[opstot['Operação'] == 'RT']
        opstot = opstot.drop(rt.index)
        rt = rt.groupby(['Operação', 'Fundo', 'Cotista', 'Datamov', 'Datacot', 'Dataliq'], as_index=False).sum()
        opstot = pd.concat([opstot, rt], join='inner')
        ops = [[list(opstot[b])[a] for b in list(opstot)] for a in range(0, len(opstot.index))]
        for m in movs_BTG:  # da pra melhorar
            m[5] = f'{transf_data(m[5])} 00:00:00'
            m[6] = f'{transf_data(m[6])} 00:00:00'
            m[7] = f'{transf_data(m[7])} 00:00:00'
            m[8] = f'{transf_data(m[8])}'
            m[9] = f'{transf_data(m[9])}'
            m[10] = f'{transf_data(m[10])}'
        for m in movs_orama:
            m[5] = f'{transf_data(m[5])} 00:00:00'
            m[6] = f'{transf_data(m[6])} 00:00:00'
            m[7] = f'{transf_data(m[7])} 00:00:00'
            m[8] = f'{transf_data(m[8])}'
            m[9] = f'{transf_data(m[9])}'
            m[10] = f'{transf_data(m[10])}'
        with xlsxwriter.Workbook(
                rf'C:\Users\{caminho_Empresa}\ROTINAS\Movimentação de Passivo\Arquivos - Excel\{data}\BTG - Empresa.xlsx') as tab:
            a = tab.add_worksheet()
            row, col = 0, 0
            for k in movs_BTG:
                a.write_row(row, col, k)
                row += 1
        with xlsxwriter.Workbook(
                        rf'C:\Users\{caminho_Empresa}\ROTINAS\Movimentação de Passivo\Arquivos - Excel\{data}\Orama - Empresa.xlsx') as tab:
            a = tab.add_worksheet()
            row, col = 0, 0
            for k in movs_orama:
                a.write_row(row, col, k)
                row += 1
        wb = openpyxl.load_workbook(
            rf'C:\Users\{caminho_Empresa}\ROTINAS\Movimentação de Passivo\Template - Mov Passivo.xlsx')
        ws = wb.sheetnames
        planilha = wb[ws[0]]
        s = 2
        for f in ops:  # [operação, valor bruto, fundo resgatado, cotista, qtdcota, datamov, dataliq]
            if f[0] not in ['TR']:
                if f[0] == 'A':
                    tp = ''
                else:
                    tp = 4
                codigo = f[3]
                for g in ['N', 'O', 'R', 'M', 'B', 'T', '-']:
                    codigo = codigo.replace(g, '')
                planilha.cell(row=s, column=1).value = int(codigo)  # id do cotista
                planilha.cell(row=s, column=2).value = int(mudancas[f[2]])
                planilha.cell(row=s, column=3).value = f[5]
                planilha.cell(row=s, column=4).value = f[6]
                planilha.cell(row=s, column=5).value = f[7]
                planilha.cell(row=s, column=6).value = mudancas[f[0]]
                planilha.cell(row=s, column=7).value = tp
                planilha.cell(row=s, column=8).value = float(f[4])
                planilha.cell(row=s, column=9).value = float(f[1]) / float(f[4])
                planilha.cell(row=s, column=10).value = f[1]
                planilha.cell(row=s, column=11).value = f[1]
                planilha.cell(row=s, column=12).value = 0
                planilha.cell(row=s, column=13).value = 0
                planilha.cell(row=s, column=14).value = 0
                planilha.cell(row=s, column=15).value = 0
                planilha.cell(row=s, column=16).value = 'Brasil'
                s += 1
            wb.save(
                rf'C:\Users\{caminho_Empresa}\ROTINAS\Movimentação de Passivo\Upload - Britech\{data}.xlsx')
        if len(movs_orama) > 0:
            outlook = win32.Dispatch('outlook.application')
            mail = outlook.CreateItem(0)
            mail.To = 'mail1'
            mail.CC = 'mail2'
            mail.Subject = f'Movimentação - Fundos - {dia}/{mes}/{ano}'
            mail.Body = 'Boa tarde! Segue anexa a movimentação de hoje!'
            attachment1 = rf'C:\Users\{caminho_Empresa}\ROTINAS\Movimentação de Passivo\Arquivos - Excel\{data}\Orama - Empresa.xlsx'
            mail.Attachments.Add(attachment1)
            mail.Display(True)
        if len(movs_BTG) > 0:
            outlook = win32.Dispatch('outlook.application')
            mail = outlook.CreateItem(0)
            mail.To = 'mail1'
            mail.CC = 'mail2'
            mail.Subject = f'Movimentação - Fundos - {dia}/{mes}/{ano}'
            mail.Body = 'Boa tarde! Segue anexa a movimentação de hoje!'
            attachment1 = rf'C:\Users\{caminho_Empresa}\ROTINAS\Movimentação de Passivo\Arquivos - Excel\{data}\BTG - Empresa.xlsx'
            mail.Attachments.Add(attachment1)
            mail.Display(True)
        net1 = 0
        net2 = 0
        tabela1 = []
        tabela2 = []
        tabela = []  # [operação, valor bruto, fundo resgatado, cotista, qtdcota, datamov, dataliq]
        for i in ops:
            tabela_ = []
            tabela_.append(i[3])
            if i[0] == 'A':
                tabela_.append(float(i[1]))
            else:
                tabela_.append(float(i[1] * -1))
            if i[2] == 'cod1':
                tabela1.append(tabela_)
            elif i[2] == 'cod2':
                tabela2.append(tabela_)
        for i in set([p[0] for p in tabela1]):
            soma = 0
            for k in tabela1:
                if k[0] == i:
                    soma += k[1]
            if soma > 0:
                tabela.append(['Aplicação', i, 'R$  {:,.2f}'.format(round(soma, 2)), '-'])
            else:
                tabela.append(['Resgate', i, 'R$  {:,.2f}'.format(round(soma, 2)), '-'])
            net1 += soma
        for i in set([p[0] for p in tabela2]):
            soma = 0
            for k in tabela2:
                if k[0] == i:
                    soma += k[1]
            if soma > 0:
                tabela.append(['Aplicação', i, '-', 'R$  {:,.2f}'.format(round(soma, 2))])
            else:
                tabela.append(['Resgate', i, '-', 'R$  {:,.2f}'.format(round(soma, 2))])
            net2 += soma
        tabela.append(
            ['Fluxo final do dia:', '-', 'R$  {:,.2f}'.format(round(net1, 2)), 'R$  {:,.2f}'.format(round(net2, 2))])
        wb = openpyxl.load_workbook(
            rf'C:\Users\{caminho_Empresa}\ROTINAS\Movimentação de Passivo\Arquivos - Excel\Template.xlsx')
        ws = wb.sheetnames
        planilha = wb[ws[0]]
        ct = 3
        for k in tabela:
            planilha.cell(row=ct, column=2).value = k[0]
            planilha.cell(row=ct, column=3).value = k[1]
            planilha.cell(row=ct, column=4).value = k[2]
            planilha.cell(row=ct, column=5).value = k[3]
            ct += 1
        wb.save(
            rf'C:\Users\{caminho_Empresa}\ROTINAS\Movimentação de Passivo\Arquivos - Excel\{data}\Tabela - Gestores.xlsx')
        excel = win32.Dispatch('Excel.Application')
        wbb = excel.Workbooks.Open(
            rf'C:\Users\{caminho_Empresa}\ROTINAS\Movimentação de Passivo\Arquivos - Excel\{data}\Tabela - Gestores.xlsx')
        sheet = wbb.Sheets[0]
        excel.visible = 1
        copyrange = sheet.Range(f'B2:E{len(tabela) + 2}')
        copyrange.CopyPicture(Appearance=1, Format=2)
        ImageGrab.grabclipboard().save(
            rf'C:\Users\{caminho_Empresa}\ROTINAS\Movimentação de Passivo\Arquivos - Excel\{data}\Tabela - Gestores.png')
        excel.Quit()

        html_body = '''
            <div>
                Prezados, boa tarde! \n \nSeguem as movimentações de hoje:
            </div>
        '''
        outlook = win32.Dispatch('outlook.application')
        mail = outlook.CreateItem(0)
        mail.To = 'mail1'
        mail.CC = 'mail2'
        mail.Subject = f'Movimentação - Fundos - {dia}/{mes}/{ano}'
        mail.HTMLBody = html_body + rf'<br><img src="C:\Users\{caminho_Empresa}\ROTINAS\Movimentação de Passivo\Arquivos - Excel\{data}\Tabela - Gestores.png">'
        mail.Display()
    tk.messagebox.showinfo('Mov. Passivo - E-mails', 'Processo finalizado! Faça o Upload na Britech agora.')


def comando4():
    ativos = []
    try:
        ativos_off = []
        caminhos = [os.path.join(
            rf'C:\Users\{caminho_Empresa}\ROTINAS\Boletagens RV\Recebidos- Trade Recaps',
            nome) for nome in os.listdir(
            rf'C:\Users\{caminho_Empresa}\ROTINAS\Boletagens RV\Recebidos- Trade Recaps')]
        movimentos = [arq for arq in caminhos if os.path.isfile(arq)]
        for i in movimentos:
            if data in i:
                with open(i, 'r') as recap:
                    ops = recap.readlines()
                    for k in ops:
                        ativo = k.split(';', 1000)
                        data_liq = ativo[1]
                        if ativo[3] == "Buy":
                            op = 'C'
                        else:
                            op = 'V'
                        ticker = f'{ativo[4]} US'
                        qtd = ativo[5]
                        preco = ativo[6]
                        cor = ativo[8]
                        if ativo[2] == 'QXP046493':
                            fd = 'cod4'
                        else:
                            fd = 'cod3'
                        at = [data, op, ticker, qtd, preco, cor, fd, data_liq, 'XP_INV']
                        ativos_off.append(at)
                        ativos.append(at)
        if len(ativos_off) != 0:
            with open(
                    fr'C:\Users\{caminho_Empresa}\ROTINAS\Boletagens RV\Enviados - Boletas\{data}.txt',
                    'w') as boleta:
                boleta.write('0#RV\n')
                for ativo in ativos_off:
                    boleta.write(
                        f'#{data}#{ativo[6]}#{ativo[1]}#N#{ativo[2]}#XP_INV#admin#{ativo[3]}#{ativo[4]}######I#####\n')
                boleta.write('99#RV')
    except:
        pass
    try:
        with open(
                fr'C:\Users\{caminho_Empresa}\ROTINAS\Boletagens RV\Recebidos - BBI\{data}.txt') as bbi:
            list = bbi.readlines()
            list.remove('0#RV\n')
            list.remove('99#RV')  # [data, op, ticker, qtd, preco, cor, fd, data_liq, 'XP_INV']
            for u in list:
                k = u.split('#', 100)
                ativos.append([k[1], k[3], k[5], k[8], k[9], '', k[2], '', k[6]])
    except:
        pass
    wb = openpyxl.load_workbook(
        rf'C:\Users\{caminho_Empresa}\ROTINAS\Boletagens RV\Template - RV.xlsx')
    ws = wb.sheetnames
    planilha = wb[ws[0]]
    s = 2
    for f in ativos:  # [data, op, ticker, qtd, preco, cor, fd, data_liq, 'XP_INV']
        if ' US' in f[2] and mudancas[f[6]] == 387894:
            conta = 158
            corretora = 'XP US'
        elif ' US' not in f[2] and mudancas[f[6]] == 387894:
            conta = 1
            corretora = f[8]
        elif ' US' not in f[2] and mudancas[f[6]] == 652091:
            conta = 164
            corretora = f[8]
        elif ' US' in f[2] and mudancas[f[6]] == 454168:
            conta = 161
            corretora = 'XP US'
        elif ' US' in f[2] and mudancas[f[6]] == 652091:
            conta = 321
            corretora = 'XP US'
        else:
            conta = 157
            corretora = f[8]
        if ' US' in f[2]:
            local = 'USA'
        else:
            local = 'BRASIL'
        planilha.cell(row=s, column=1).value = mudancas[f[6]]
        planilha.cell(row=s, column=2).value = f[2]
        planilha.cell(row=s, column=3).value = mudancas[corretora]
        planilha.cell(row=s, column=4).value = f[1]
        planilha.cell(row=s, column=5).value = muda_data(f[0])
        planilha.cell(row=s, column=6).value = f[4].replace('.', ',')
        planilha.cell(row=s, column=7).value = str(float(f[4].replace(',', '.')) * int(f[3])).replace('.', ',')
        planilha.cell(row=s, column=8).value = int(f[3])
        planilha.cell(row=s, column=11).value = 'S'
        planilha.cell(row=s, column=12).value = local
        planilha.cell(row=s, column=13).value = 'CBLC'
        planilha.cell(row=s, column=14).value = 'CBLC'
        planilha.cell(row=s, column=15).value = 10003
        planilha.cell(row=s, column=21).value = conta
        s += 1
    wb.save(
        rf'C:\Users\{caminho_Empresa}\ROTINAS\Boletagens RV\Upload - Britech\{data}.xlsx')
    tk.messagebox.showinfo('Boletagem - RV',
                           "A Boletagem foi feita com sucesso e os arquivos para Upload já se encontram na pasta.")


def comando5():
    dados5 = []
    prompt5 = tk.Tk()
    c5 = tk.Canvas(prompt5, width=600, height=550, relief='raised', bg=preto_Empresa)
    c5.pack()
    v51 = tk.StringVar(prompt5)
    v51.set('Selecione')
    v52 = tk.StringVar(prompt5)
    v52.set('Selecione')
    v53 = tk.StringVar(prompt5)
    v54 = tk.StringVar(prompt5)
    v55 = tk.StringVar(prompt5)
    v56 = tk.StringVar(prompt5)
    l5 = tk.Label(prompt5, text='Boletagem - Câmbio', font=('helvetica', 14), bg=preto_Empresa, fg=amarelo_Empresa)
    c5.create_window(300, 25, window=l5)

    l51 = tk.Label(prompt5, text='Insira o tipo de operação: Compra ou Venda',
                   font=('helvetica', 10), bg=preto_Empresa, fg=amarelo_Empresa)
    c5.create_window(300, 70, window=l51)
    ent51 = tk.OptionMenu(prompt5, v51, *['Compra', 'Venda'])
    c5.create_window(300, 95, window=ent51)

    l52 = tk.Label(prompt5, text='Em qual fundo foi feita a operação?',
                   font=('helvetica', 10), bg=preto_Empresa, fg=amarelo_Empresa)
    c5.create_window(300, 135, window=l52)
    ent52 = tk.OptionMenu(prompt5, v52, *['Fundo I', 'Fundo II'])
    c5.create_window(300, 160, window=ent52)

    l53 = tk.Label(prompt5, text='Qual a quantia de dólar operada?',
                   font=('helvetica', 10), bg=preto_Empresa, fg=amarelo_Empresa)
    c5.create_window(300, 200, window=l53)
    ent53 = tk.Entry(prompt5, textvariable=v53)
    c5.create_window(300, 225, window=ent53)

    l54 = tk.Label(prompt5, text='Qual a taxa de Câmbio praticado? (escreva com "." na casa decimal)',
                   font=('helvetica', 10), bg=preto_Empresa, fg=amarelo_Empresa)
    c5.create_window(300, 265, window=l54)
    ent54 = tk.Entry(prompt5, textvariable=v54)
    c5.create_window(300, 290, window=ent54)

    l55 = tk.Label(prompt5, text='Insira a data de saída do recurso. Ex: AAAAMMDD (sem as barras)',
                   font=('helvetica', 10), bg=preto_Empresa, fg=amarelo_Empresa)
    c5.create_window(300, 330, window=l55)
    ent55 = tk.Entry(prompt5, textvariable=v55)
    c5.create_window(300, 355, window=ent55)

    l56 = tk.Label(prompt5, text='Insira a data de liquidação da operação. Ex: AAAAMMDD (sem as barras)',
                   font=('helvetica', 10), bg=preto_Empresa, fg=amarelo_Empresa)
    c5.create_window(300, 395, window=l56)
    ent56 = tk.Entry(prompt5, textvariable=v56)
    c5.create_window(300, 420, window=ent56)

    def pegainput5():
        global w51, w52, w53, w54, w55, w56
        w51 = v51.get()
        w52 = v52.get()
        w53 = ent53.get()
        w54 = ent54.get()
        w55 = ent55.get()
        w56 = ent56.get()
        dados5.append([w51, w52, w53, w54, w55, w56])
        v51.set('Selecione')
        v52.set('Selecione')
        v53.set('')
        v54.set('')
        v55.set('')
        v56.set('')

    bot5 = tk.Button(prompt5, text='Registrar >', command=pegainput5, bg=branco_Empresa, fg=preto_Empresa,
                     font=('helvetica', 9, 'bold'))
    c5.create_window(300, 500, window=bot5)
    bot52 = tk.Button(prompt5, text='Concluir!', command=lambda: [f() for f in [prompt5.destroy, root1.quit]],
                      bg=amarelo_Empresa, fg=preto_Empresa,
                      font=('helvetica', 9, 'bold'))
    c5.create_window(400, 500, window=bot52)

    prompt5.mainloop()
    lista_cam = []
    for i in dados5:
        if i[1] == 'Fundo I':
            cod = 'cod4'
        else:
            cod = 'cod3'
        if i[0] == "Compra":
            lista_cam.append([cod, 'DOLAR', "REAL", float(i[2]), float(i[3]), i[4], i[5], float(i[2]) * float(i[3]), "USA", 'BRASIL'])
        else:
            lista_cam.append([cod, 'REAL', "DOLAR", float(i[2]) * float(i[3]), float(i[3]), i[4], i[5], float(i[2]), 'BRASIL', 'USA'])
    with open(
            fr'C:\Users\{caminho_Empresa}\ROTINAS\Câmbio\Upload admin\{data}.txt',
            'w') as cambio:
        cambio.write('0#CA\n')
        for k in lista_cam:
            cambio.write(
                f'#{data}#{k[0]}#V#U#D0#V##{k[1]}#admin#{k[3]}##{k[4]}#{k[8]}#{k[2]}#{k[9]}#{k[7]}#{data}##{k[6]}#{k[5]}##{k[6]}######\n')
        cambio.write('99#CA')
    tk.messagebox.showinfo('Boletagem - Câmbio',
                           "A boletagem foi feita e está disponível na pasta. \nPreencha agora as informações com os dados referentes no email na aba 'Câmbio'\nATENÇÃO: para VENDA de dólar, é necessário assinar e enviar a Wire!\nPor fim, bolete no Sistema da Britech")


def comando6():
    wb = openpyxl.load_workbook(
        fr'C:\Users\{caminho_Empresa}\ROTINAS\Aluguel\Operações - XP\{data}.xlsx')
    ws = wb.sheetnames
    planilha = wb[ws[0]]
    nops = planilha.max_row
    n = 5
    ativos = []
    dictaluguel = {'Empresa I': 'cod4', 'Empresa II': 'cod3'}
    while n <= nops:
        cpf = dictaluguel[planilha.cell(row=n, column=2).value]
        ticker = planilha.cell(row=n, column=4).value
        cor = 'XP_INV'
        op = planilha.cell(row=n, column=5).value
        qtd = int(planilha.cell(row=n, column=6).value)
        preco = float(planilha.cell(row=n, column=10).value)
        com = 0.15  # mudar pra "15,00" se conseguirmos acertar a taxa de comissão.
        rem = str(float(planilha.cell(row=n, column=7).value)).replace('.', ',')
        a = planilha.cell(row=n, column=8).value
        data_venc = f'{a[6:10]}{a[3:5]}{a[0:2]}'
        if 'alc' in planilha.cell(row=n, column=15).value:
            tipo = 'R'
            mod = ''
        else:
            tipo = "N"
            mod = 'E1'
        codigo = 1
        d = datetime.datetime(int(data[0:4]), int(data[4:6]), int(data[6:8]))
        if d.weekday() == 4 or d.weekday() == 3:
            d_2 = d + datetime.timedelta(days=4)
        else:
            d_2 = d + datetime.timedelta(days=2)
        d2 = str(d_2).replace('-', '')
        d2 = d2.replace(' 00:00:00', '')
        ativos.append([data, op, ticker, cor, data_venc, com, rem, qtd, cpf, codigo, tipo, mod, d2, preco])
        n += 1
    with open(
            fr'C:\Users\{caminho_Empresa}\ROTINAS\Aluguel\Upload Custódia\{data}.txt',
            'w') as boleta:
        boleta.write('0#EM\n')
        for ativo in ativos:
            boleta.write(
                f'#{data}#{ativo[1]}#{ativo[8]}#{ativo[2]}#{ativo[7]}#I#S#{ativo[12]}##{ativo[3]}#{ativo[5]}#{ativo[6]}#{ativo[4]}####I#####{ativo[10]}#{ativo[11]}#R##\n')
        boleta.write('99#EM')
    wb = openpyxl.load_workbook(
        rf'C:\Users\{caminho_Empresa}\ROTINAS\Aluguel\Template - Britech - Aluguel.xlsx')
    ws = wb.sheetnames
    planilha = wb[ws[0]]
    s = 2
    for f in ativos:
        planilha.cell(row=s, column=1).value = mudancas[f[8]]
        planilha.cell(row=s, column=2).value = f[2]
        planilha.cell(row=s, column=3).value = 10003  #mudancas[f[3]] <- esse comando cadastra a corretora, enquanto 10003 registra como tendo sido feito pela corretora "Banco admin"
        planilha.cell(row=s, column=4).value = mudancas[f[1]]
        planilha.cell(row=s, column=5).value = muda_data(f[0])
        planilha.cell(row=s, column=6).value = muda_data(f[4])
        planilha.cell(row=s, column=7).value = f[6]  # caso a importação de taxa de remuneração não funcionar, é str(float(f[6].replace(',', '.'))*(1-f[5])).replace('.', ',')
        planilha.cell(row=s, column=8).value = f[7]
        planilha.cell(row=s, column=9).value = f[13]
        planilha.cell(row=s, column=10).value = f[9]
        planilha.cell(row=s, column=11).value = 1
        planilha.cell(row=s, column=13).value = str(float(f[6].replace(',', '.'))*f[5]).replace('.', ',')# é 0 se não funcionar a importação de taxa de remuneração
        planilha.cell(row=s, column=14).value = 'S'
        planilha.cell(row=s, column=15).value = muda_data(f[0])
        planilha.cell(row=s, column=16).value = muda_data(f[0])
        planilha.cell(row=s, column=18).value = muda_data(f[0])
        s += 1
    wb.save(
        rf'C:\Users\{caminho_Empresa}\ROTINAS\Aluguel\Upload - Britech\{data}.xlsx')
    tk.messagebox.showinfo('Boletagem - Aluguel',
                           "A Boletagem foi feita com sucesso e os arquivos para Upload já se encontram na pasta.")


def comando7():
    dados7 = []
    prompt7 = tk.Tk()
    c7 = tk.Canvas(prompt7, width=500, height=350, relief='raised', bg=preto_Empresa)
    c7.pack()
    v71 = tk.StringVar(prompt7)
    v71.set('Selecione')
    v72 = tk.StringVar(prompt7)
    v72.set('Selecione')
    v73 = tk.StringVar(prompt7)
    l7 = tk.Label(prompt7, text='Geração de Caixa', font=('helvetica', 14), bg=preto_Empresa, fg=amarelo_Empresa)
    c7.create_window(250, 25, window=l7)

    l71 = tk.Label(prompt7, text='A operação foi feita no Fundo I ou Fundo II?',
                   font=('helvetica', 10), bg=preto_Empresa, fg=amarelo_Empresa)
    c7.create_window(250, 70, window=l71)
    ent71 = tk.OptionMenu(prompt7, v71, *['Fundo I', 'Fundo II'])
    c7.create_window(250, 95, window=ent71)

    l72 = tk.Label(prompt7, text="A operação foi é Aplicação ou Resgate?",
                   font=('helvetica', 10), bg=preto_Empresa, fg=amarelo_Empresa)
    c7.create_window(250, 135, window=l72)
    ent72 = tk.OptionMenu(prompt7, v72, *['Aplicação', 'Resgate'])
    c7.create_window(250, 160, window=ent72)

    l73 = tk.Label(prompt7, text='Qual a quantia resgatada/aplicada?',
                   font=('helvetica', 10), bg=preto_Empresa, fg=amarelo_Empresa)
    c7.create_window(250, 200, window=l73)
    ent73 = tk.Entry(prompt7, textvariable=v73)
    c7.create_window(250, 225, window=ent73)

    def pegainput7():
        global w71, w72, w73
        w71 = v71.get()
        w72 = v72.get()
        w73 = ent73.get()
        dados7.append([w71, w72, w73])
        v71.set('Selecione')
        v72.set('Selecione')
        v73.set('')

    bot7 = tk.Button(prompt7, text='Registrar >', command=pegainput7, bg=branco_Empresa, fg=preto_Empresa,
                     font=('helvetica', 9, 'bold'))
    c7.create_window(250, 300, window=bot7)
    bot72 = tk.Button(prompt7, text='Prosseguir >>>', command=lambda: [f() for f in [prompt7.destroy, root1.quit]],
                      bg=amarelo_Empresa, fg=preto_Empresa,
                      font=('helvetica', 9, 'bold'))
    c7.create_window(350, 300, window=bot72)
    prompt7.mainloop()

    ops = []
    for i in dados7:
        if i[1] == 'Aplicação':
            ope = 'A'
        else:
            ope = 'F'
        if i[0] == 'Empresa I':
            ops.append([ope, float(i[2]), 'cod4', 'cod1', '20352006'])
        else:
            ops.append([ope, float(i[2]), 'cod3', 'cod2', '26587002'])
    with open(
            fr'C:\Users\{caminho_Empresa}\ROTINAS\Resgates - Caixa\Upload - Custódia\{data} - Ponta ATIVO.txt',
            'w') as boleta:
        boleta.write('0#FD\n')
        for ativo in ops:
            boleta.write(f'#{data}#{ativo[3]}#{ativo[0]}#{ativo[2]}###{ativo[1]}###admin#CETIP#2#############\n')
        boleta.write('99#FD')
    lista_mov = []
    for j in ops:
        if j[0] == "F":
            op = 'RB'
        else:
            op = 'A'
        if j[2] == "cod4":
            fd = 'EMP0030'
            cc = '13717'
        else:
            fd = f'EMP0012'
            cc = '20466'
        if j[0] == "F":
            brut = 'I'
        else:
            brut = ''
        lista_mov.append(
            ['', op, 'N', fd, j[2], f'{dia}/{mes}/{ano}',
             str(j[1]).replace('.', ','), '', '', 'CC', '237', '2856', cc, brut, 'R', 'STR', 'N', '', '', '', '', '', '', 'N', ''])
    with open(
            rf'C:\Users\{caminho_Empresa}\ROTINAS\Resgates - Caixa\Upload - Custódia\{data} - Ponta PASSIVO.txt',
            'w') as upload:
        for i in lista_mov:
            for j in i:
                upload.write(f'{j}\t')
            upload.write('\n')
    tk.messagebox.showinfo('Gerar Caixa',
                           'Os arquivos foram feitos e se encontram na pasta "Resgates - Caixa"\nFaça o upload deles nas áreas respectivas.')


def comando8():
    dados8 = []
    prompt8 = tk.Tk()
    c8 = tk.Canvas(prompt8, width=500, height=350, relief='raised', bg=preto_Empresa)
    c8.pack()
    v81 = tk.StringVar(prompt8)
    v82 = tk.StringVar(prompt8)
    v83 = tk.StringVar(prompt8)
    l8 = tk.Label(prompt8, text='Cadastro de Cotistas', font=('helvetica', 14), bg=preto_Empresa, fg=amarelo_Empresa)
    c8.create_window(250, 25, window=l8)

    l81 = tk.Label(prompt8, text='Digite o número do distribuidor para o qual deseja fazer o cadastro?',
                   font=('helvetica', 10), bg=preto_Empresa, fg=amarelo_Empresa)
    c8.create_window(250, 70, window=l81)
    ent81 = tk.OptionMenu(prompt8, v81, *['dist', 'BTG'])
    c8.create_window(250, 95, window=ent81)

    l82 = tk.Label(prompt8, text='O cadastro é para qual fundo?',
                   font=('helvetica', 10), bg=preto_Empresa, fg=amarelo_Empresa)
    c8.create_window(250, 135, window=l82)
    ent82 = tk.OptionMenu(prompt8, v82, *['Empresa I', 'Empresa II'])
    c8.create_window(250, 160, window=ent82)

    l83 = tk.Label(prompt8, text='Qual o código do cotista?',
                   font=('helvetica', 10), bg=preto_Empresa, fg=amarelo_Empresa)
    c8.create_window(250, 200, window=l83)
    ent83 = tk.Entry(prompt8, textvariable=v83)
    c8.create_window(250, 225, window=ent83)

    def pegainput8():
        global w81, w82, w83
        w81 = v81.get()
        w82 = v82.get()
        w83 = ent83.get()
        dados8.append([w81, w82, w83])
        v81.set('Selecione')
        v81.set('Selecione')
        v81.set('')

    bot8 = tk.Button(prompt8, text='Registrar >', command=pegainput8, bg=branco_Empresa, fg=preto_Empresa,
                     font=('helvetica', 9, 'bold'))
    c8.create_window(250, 300, window=bot8)
    bot82 = tk.Button(prompt8, text='Prosseguir >>>', command=lambda: [f() for f in [prompt8.destroy, root1.quit]],
                      bg=amarelo_Empresa, fg=preto_Empresa,
                      font=('helvetica', 9, 'bold'))
    c8.create_window(350, 300, window=bot82)
    prompt8.mainloop()
    orama = []
    btg = []
    for i in dados8:
        corretora = int(i[0])
        if corretora == 'dist':
            orama.append([int(i[1]), i[2]])
            wbo = openpyxl.load_workbook(
                rf'C:\Users\{caminho_Empresa}\ROTINAS\Cadastro de Cotistas por Conta e Ordem\Template.xlsx')
            wbo.save(
                rf'C:\Users\{caminho_Empresa}\ROTINAS\Cadastro de Cotistas por Conta e Ordem\dist\Cadastro Cotistas - {dia}/{mes}/{ano}.xlsx')
        elif corretora == 'BTG':
            btg.append([int(i[1]), i[2]])
            wbb = openpyxl.load_workbook(
                rf'C:\Users\{caminho_Empresa}\ROTINAS\Cadastro de Cotistas por Conta e Ordem\Template.xlsx')
            wbb.save(
                rf'C:\Users\{caminho_Empresa}\ROTINAS\Cadastro de Cotistas por Conta e Ordem\BTG\Cadastro Cotistas - {dia}/{mes}/{ano}.xlsx')
    for k in orama:
        fundo = k[0]
        codigo = k[1]
        qt_orama = 3
        a = 'EMP-'
        mae = ''.join(char for char in codigo if char not in a)
        wb = openpyxl.load_workbook(
            rf'C:\Users\{caminho_Empresa}\ROTINAS\Cadastro de Cotistas por Conta e Ordem\dist\Cadastro Cotistas - {dia}/{mes}/{ano}.xlsx')
        ws = wb.sheetnames
        planilha = wb[ws[0]]
        planilha.cell(row=qt_orama, column=1).value = f'{mae}'
        planilha.cell(row=qt_orama, column=2).value = f'{codigo}'
        planilha.cell(row=qt_orama, column=3).value = 'ORAMA DTVM'
        planilha.cell(row=qt_orama, column=4).value = '13.293.225/0001-25'
        planilha.cell(row=qt_orama, column=5).value = 'Itaú'
        planilha.cell(row=qt_orama, column=6).value = '0272'
        planilha.cell(row=qt_orama, column=7).value = '13002-2'
        planilha.cell(row=qt_orama, column=8).value = 'admin'
        if fundo == 'Empresa I':
            planilha.cell(row=qt_orama, column=9).value = '2373'
            planilha.cell(row=qt_orama, column=10).value = '32108-7'
            planilha.cell(row=qt_orama, column=14).value = 'cod1'
            planilha.cell(row=qt_orama, column=15).value = 'cnpj1'
            planilha.cell(row=qt_orama, column=16).value = 'Empresa I FIC FIA'
        else:
            planilha.cell(row=qt_orama, column=9).value = '2373'
            planilha.cell(row=qt_orama, column=10).value = '32108-7'
            planilha.cell(row=qt_orama, column=14).value = 'cod2'
            planilha.cell(row=qt_orama, column=15).value = 'cnpj2'
            planilha.cell(row=qt_orama, column=16).value = 'Empresa II FIC FIA'
        qt_orama += 1
        wb.save(
            rf'C:\Users\{caminho_Empresa}\ROTINAS\Cadastro de Cotistas por Conta e Ordem\dist\Cadastro Cotistas - {dia}/{mes}/{ano}.xlsx')
        outlook = win32.Dispatch('outlook.application')
        mail = outlook.CreateItem(0)
        mail.To = 'mail1'
        mail.CC = 'mail2'
        mail.Subject = 'Cadastro de Cotistas por Conta e Ordem - Empresa'
        mail.Body = 'Prezados, boa tarde! \n \n Segue planilha com os dados de novos cotistas para cadastro do fundo. \n \n Favor informar quando o cadastro estiver finalizado \n \n Att,'
        attachment1 = rf'C:\Users\{caminho_Empresa}\ROTINAS\Cadastro de Cotistas por Conta e Ordem\dist\Cadastro Cotistas - {dia}/{mes}/{ano}.xlsx'
        mail.Attachments.Add(attachment1)
        mail.Display()

    tk.messagebox.showinfo('Cadastro de Cotista',
                           'O processo foi finalizado com sucesso.')


def comando9():
    jorge = pd.read_excel(fr'C:\Users\{caminho_Empresa}\ROTINAS\Relatório de Cotações\Base.xlsx', 0)
    esse_ano = jorge.tail(252).set_index('Data Referência')
    lista_cotacoes = []
    for i in [a for a in list(esse_ano) if a != 'Data Referência']:
        a = esse_ano[i]
        hj = a.get(key=pd.to_datetime(f'{ontem}'))
        try:
            ult_mes = a.get(key=pd.to_datetime(f'{ontem}') - datetime.timedelta(int(dia)-1))
            d = 1/ult_mes
        except:
            try:
                ult_mes = a.get(key=pd.to_datetime(f'{ontem}') - datetime.timedelta(int(dia)))
                d = 1 / ult_mes
            except:
                ult_mes = a.get(key=pd.to_datetime(f'{ontem}') - datetime.timedelta(int(dia) + 1))
        try:
            tresm = a.get(key=pd.to_datetime(f'{ontem}') - datetime.timedelta(60))
            d = 1 / tresm
        except:
            try:
                tresm = a.get(key=pd.to_datetime(f'{ontem}') - datetime.timedelta(61))
                d = 1 / tresm # linhas para testar a não nulidade do valor
            except:
                tresm = a.get(key=pd.to_datetime(f'{ontem}') - datetime.timedelta(62))
        um_ano = a.get(key=0)
        ytd = a.get(key='2022-01-03')
        max = a.max()
        min = a.min()
        lista_cotacoes.append(
            [i, (hj / ult_mes - 1), (hj / tresm - 1), (hj / ytd - 1), (hj / um_ano - 1), (hj / min - 1), (hj/max - 1)])
    tab = openpyxl.open(
        fr'C:\Users\{caminho_Empresa}\ROTINAS\Relatório de Cotações\Template - Relatório.xlsx')
    ws = tab.sheetnames
    planilha = tab[ws[1]]
    r = 0
    for f in lista_cotacoes:
        planilha.cell(row=2 + r, column=3).value = f[0]
        planilha.cell(row=2 + r, column=4).value = f[1]
        planilha.cell(row=2 + r, column=5).value = f[2]
        planilha.cell(row=2 + r, column=6).value = f[3]
        planilha.cell(row=2 + r, column=7).value = f[4]
        planilha.cell(row=2 + r, column=8).value = f[5]
        planilha.cell(row=2 + r, column=9).value = f[6]
        r += 1
    tab.save(fr'C:\Users\{caminho_Empresa}\ROTINAS\Relatório de Cotações\Enviados\Relatório de Cotações - {dia}-{mes}-{ano}.xlsx')
    outlook = win32.Dispatch('outlook.application')
    mail = outlook.CreateItem(0)
    mail.To = 'mail1'
    mail.CC = 'mail2'
    mail.Subject = f'Relatório Cotações - {dia}/{mes}/{ano}'
    mail.Body = f'Bom dia! \n\nSegue anexa a planilha com o relatório de Cotações relativo ao dia de hoje.\n\nAtt,'
    mail.Attachments.Add(
        fr'C:\Users\{caminho_Empresa}\ROTINAS\Relatório de Cotações\Enviados\Relatório de Cotações - {dia}-{mes}-{ano}.xlsx')
    mail.Display(True)


def comando10():
    caminhos = [os.path.join(
        rf'C:\Users\{caminho_Empresa}\ROTINAS\Relatório de Cotações\Novos ativos',
        nome) for nome in os.listdir(
        rf'C:\Users\{caminho_Empresa}\ROTINAS\Relatório de Cotações\Novos ativos')]
    movimentos = [arq for arq in caminhos if os.path.isfile(arq)]
    if len(movimentos) >= 1:
        jorge = pd.read_excel(
            fr'C:\Users\{caminho_Empresa}\ROTINAS\Relatório de Cotações\Base.xlsx', 0)
        for i in movimentos:
            f = pd.read_excel(i, 0)
            f = f.drop(columns=['PU Médio', 'PU Abertura'])
            g = pd.pivot_table(f, values='Código', index='Data Referência', columns='Código').reset_index()
            jorge = jorge.merge(g, on='Data Referência', how='left')
        jorge.to_excel(fr'C:\Users\{caminho_Empresa}\ROTINAS\Relatório de Cotações\Base.xlsx',
                       index=False)
    else:
        pass
    # Atualiza a base com o arquivão que tem todas as cotações (de download diário)
    try:
        jorge = pd.read_excel(fr'C:\Users\{caminho_Empresa}\ROTINAS\Relatório de Cotações\Base.xlsx', 0)
        jorge = jorge.set_index('Data Referência')
        novo = pd.read_excel(fr'C:\Users\{caminho_Empresa}\ROTINAS\Relatório de Cotações\Histórico - Britech\{ontem}.xlsx', 0)
        novo = novo.drop(columns=['PU Médio', 'PU Abertura'])
        novo = pd.pivot_table(novo, values='Código',index='Data Referência', columns='Código').reset_index()
        novo = novo.set_index('Data Referência')
        jorge = pd.concat([jorge, novo], join='inner').reset_index()
        jorge.to_excel(fr'C:\Users\{caminho_Empresa}\ROTINAS\Relatório de Cotações\Base.xlsx', index=False)
    except:
        pass
    tk.messagebox.showinfo('Relatório de Cotações - Alimenta base',
                           'A base foi alimentada com sucesso!')


def comando11():
    caminhos = [os.path.join(
        rf'C:\Users\{caminho_Empresa}\ROTINAS\PASTA INPUT',
        nome) for nome in os.listdir(
        rf'C:\Users\{caminho_Empresa}\ROTINAS\PASTA INPUT')]
    movimentos = [arq for arq in caminhos if os.path.isfile(arq)]
    for i in movimentos:
        if ontem in i and 'Empresa I MASTER' in i and '.XML' in i:
            os.rename(i,
                      fr'C:\Users\{caminho_Empresa}\ROTINAS\Balanco\Carteiras XML\Empresa I MASTER\{ontem}.xml')
            shutil.copyfile(
                fr'C:\Users\{caminho_Empresa}\Empresa I MASTER\{ontem}.xml',
                fr'C:\Users\{caminho_Empresa}\Empresa I Master\{ontem[0:4]}.{ontem[4:6]}.{ontem[6:8]}_Master_FIA.xml')
        elif ontem in i and 'Empresa I FIC' in i and '.XML' in i:
            os.rename(i,
                      fr'C:\Users\{caminho_Empresa}\ROTINAS\Balanco\Carteiras XML\Empresa I FIC\{ontem}.xml')
            shutil.copyfile(
                fr'C:\Users\{caminho_Empresa}\ROTINAS\Balanco\Carteiras XML\Empresa I FIC\{ontem}.xml',
                fr'C:\Users\{caminho_Empresa}\Empresa I FIC\{ontem[0:4]}.{ontem[4:6]}.{ontem[6:8]}.xml')
        elif ontem in i and 'Empresa II FIC FIA' in i and '.XML' in i:
            os.rename(i,
                      fr'C:\Users\{caminho_Empresa}\ROTINAS\Balanco\Carteiras XML\Empresa II FIC FIA\{ontem}.xml')
            shutil.copyfile(
                fr'C:\Users\{caminho_Empresa}\ROTINAS\Balanco\Carteiras XML\Empresa II FIC FIA\{ontem}.xml',
                fr'C:\Users\{caminho_Empresa}\Empresa II FIC\{ontem[0:4]}.{ontem[4:6]}.{ontem[6:8]} FIC FIA.xml')
        elif ontem in i and 'Empresa II MASTER FIA' in i and '.XML' in i:
            os.rename(i,
                      fr'C:\Users\{caminho_Empresa}\ROTINAS\Balanco\Carteiras XML\Empresa II MASTER FIA\{ontem}.xml')
            shutil.copyfile(
                fr'C:\Users\{caminho_Empresa}\ROTINAS\Balanco\Carteiras XML\Empresa II MASTER FIA\{ontem}.xml',
                fr'C:\Users\{caminho_Empresa}\Empresa II Master\{ontem[0:4]}.{ontem[4:6]}.{ontem[6:8]} MASTER FIA.xml')
        elif f'Empresa_{data}' in i:
            os.rename(i,
                      fr'C:\Users\{caminho_Empresa}\ROTINAS\Boletagens RV\Recebidos- Trade Recaps\{data}.txt')
        elif f'Ranking Diário_{ontem_n_tao_bonito}' in i:
            os.rename(i,
                      fr'C:\Users\{caminho_Empresa}\ROTINAS\Relatório - BTC\Recebidos - BTG\Ranking Diário_{ontem_n_tao_bonito}.xlsm')
        elif 'Aluguel' in i and f'{dia}{mes}{ano}' in i:
            os.rename(i,
                      fr'C:\Users\{caminho_Empresa}\ROTINAS\Relatório - BTC\Posições BTC - BBI\{ontem}.csv')
        elif 'Garantias' in i and f'{dia}{mes}{ano}' in i:
            os.rename(i,
                      fr'C:\Users\{caminho_Empresa}\ROTINAS\Relatório - BTC\Garantias - BBI\{ontem}.csv')
        elif 'RV_EMP' in i:
            os.rename(i,
                      fr'C:\Users\{caminho_Empresa}\ROTINAS\Boletagens RV\Recebidos - BBI\{data}.txt')
        elif 'FINA' in i:
            os.rename(i,
                      fr'C:\Users\{caminho_Empresa}\ROTINAS\Balanco\FINAs\FINA_{dia}{mes}.xml')
        elif 'Movimentos (' in i and '.xls' in i:
            os.rename(i,
                      fr'C:\Users\{caminho_Empresa}\ROTINAS\Balanco\Movimentos\Posições D-30\{data}.xls')
        elif 'BTG - Movimentacoes' and ' FIA' in i:
            try:
                os.rename(i,
                          fr'C:\Users\{caminho_Empresa}\ROTINAS\Movimentação de Passivo\Arquivos - Recebidos\{data}\BTG - Empresa I.xls')
            except:
                os.mkdir(
                    fr'C:\Users\{caminho_Empresa}\ROTINAS\Movimentação de Passivo\Arquivos - Recebidos\{data}')
                os.rename(i,
                          fr'C:\Users\{caminho_Empresa}\ROTINAS\Movimentação de Passivo\Arquivos - Recebidos\{data}\BTG - Empresa I.xls')
        elif 'BTG - Movimentacoes' and 'FICFIA' in i:
            try:
                os.rename(i,
                          fr'C:\Users\{caminho_Empresa}\ROTINAS\Movimentação de Passivo\Arquivos - Recebidos\{data}\BTG - Empresa II.xls')
            except:
                os.mkdir(
                    fr'C:\Users\{caminho_Empresa}\ROTINAS\Movimentação de Passivo\Arquivos - Recebidos\{data}')
                os.rename(i,
                          fr'C:\Users\{caminho_Empresa}\ROTINAS\Movimentação de Passivo\Arquivos - Recebidos\{data}\BTG - Empresa II.xls')
        elif 'Orama_' in i and 'II-FIC-FIA' in i:
            try:
                os.rename(i,
                          fr'C:\Users\{caminho_Empresa}\ROTINAS\Movimentação de Passivo\Arquivos - Recebidos\{data}\Orama - Empresa II.xls')
            except:
                os.mkdir(
                    fr'C:\Users\{caminho_Empresa}\ROTINAS\Movimentação de Passivo\Arquivos - Recebidos\{data}')
                os.rename(i,
                          fr'C:\Users\{caminho_Empresa}\ROTINAS\Movimentação de Passivo\Arquivos - Recebidos\{data}\Orama Empresa II.xls')
        elif 'Orama_' in i and 'I-FIC-FIA' in i:
            try:
                os.rename(i,
                          fr'C:\Users\{caminho_Empresa}\ROTINAS\Movimentação de Passivo\Arquivos - Recebidos\{data}\Orama Empresa I.xls')
            except:
                os.mkdir(
                    fr'C:\Users\{caminho_Empresa}\ROTINAS\Movimentação de Passivo\Arquivos - Recebidos\{data}')
                os.rename(i,
                          fr'C:\Users\{caminho_Empresa}\ROTINAS\Movimentação de Passivo\Arquivos - Recebidos\{data}\Orama Empresa I.xls')
        elif 'Movimentos_20' in i:
            os.mkdir(
                fr'C:\Users\{caminho_Empresa}\ROTINAS\Movimentação de Passivo\Arquivos - Excel\{data}')
            os.rename(i,
                      fr'C:\Users\{caminho_Empresa}\ROTINAS\Movimentação de Passivo\Arquivos - Excel\{data}\Movimentos - admin.xls')
        elif 'taxa_selic' in i:
            os.rename(i,
                      fr'C:\Users\{caminho_Empresa}\ROTINAS\Relatorios Diarios\Bases de dados\Arquivos CDI\{ontem}.txt')
        elif 'EMPRESA CTA' in i or 'Empresa CTA' in i:
            os.rename(i,
                      fr'C:\Users\{caminho_Empresa}\ROTINAS\Aluguel\Operações - XP\{data}.xlsx')
        elif 'ContractLegs' in i and f'{ano}{mes}{dia}' in i:
            os.rename(i,
                      fr'C:\Users\{caminho_Empresa}\ROTINAS\Relatório - BTC\Posições BTC - BIP\{ontem}.csv')
        elif 'CollateralValuation' in i and f'{ano}{mes}{dia}' in i:
            os.rename(i,
                      fr'C:\Users\{caminho_Empresa}\ROTINAS\Relatório - BTC\Garantias - BIP\{ontem}.csv')
        elif f'.zip' in i:
            os.mkdir(
                fr'C:\Users\{caminho_Empresa}\ROTINAS\Zeragem - Britech\Operacoes admin\{ontem}')
            os.rename(i,
                      fr'C:\Users\{caminho_Empresa}\ROTINAS\Zeragem - Britech\Operacoes admin\{ontem}\ZIP.zip')
            with ZipFile(fr'C:\Users\{caminho_Empresa}\ROTINAS\Zeragem - Britech\Operacoes admin\{ontem}\ZIP.zip', 'r') as zip:
                zip.extractall(path=fr'C:\Users\{caminho_Empresa}\ROTINAS\Zeragem - Britech\Operacoes admin\{ontem}')
        elif 'gridCadastro (' in i:
            os.rename(i,
                      fr'C:\Users\{caminho_Empresa}\ROTINAS\Relatório de Cotações\Histórico - Britech\{ontem}.xlsx')


def comando12():
    btcs = cart_m1_ont.btcs()
    gara = cart_m1_ont.garantias()
    try:
        dataframebtg_brut = pd.read_excel(
            fr'C:\Users\{caminho_Empresa}\ROTINAS\Relatório - BTC\Recebidos - BTG\Ranking Diário_{ontem_n_tao_bonito}.xlsm',
            sheet_name=1)
    except:
        dataframebtg_brut = pd.read_excel(
            fr'C:\Users\{caminho_Empresa}\ROTINAS\Relatório - BTC\Recebidos - BTG\Ranking Diário_{muda_data(anteontem).replace("/","")}.xlsm',
            sheet_name=1)
    dataframebtg = pd.DataFrame()
    dataframebtg['Ticker'] = dataframebtg_brut['Unnamed: 0']
    dataframebtg['Taxa média'] = dataframebtg_brut['Unnamed: 6']
    dataframebtg['SIR'] = dataframebtg_brut['Unnamed: 8']
    dataframeemp_brut1 = pd.DataFrame()
    dataframeemp_brut2 = pd.DataFrame()
    dataframeemp_brut1['Ticker'] = [a[0] for a in cart_m1_ont.acoes]
    dataframeemp_brut1['QTD comprada Empresa I'] = [a[1] if a[6] in ['C', 'D'] else 0 for a in cart_m1_ont.acoes]
    dataframeemp_brut1['QTD vendida Empresa I'] = [-1 * a[1] if a[6] == 'V' else 0 for a in cart_m1_ont.acoes]
    dataframeemp_brut1['QTD bloq. Empresa I'] = [a[4] for a in cart_m1_ont.acoes]
    dataframeemp_brut1['QTD doada Empresa I'] = [a[4] if a[6] == 'D' else 0 for a in cart_m1_ont.acoes]
    dataframeemp_brut1['QTD tomada Empresa I'] = [a[4] if a[6] == 'T' else 0 for a in cart_m1_ont.acoes]
    dataframeemp_brut1['QTD vendida Empresa I'] = dataframeemp_brut1['QTD tomada Empresa I'] - dataframeemp_brut1[
        'QTD vendida Empresa I']
    dataframeemp_brut1['QTD livre Empresa I'] = dataframeemp_brut1['QTD comprada Empresa I'] + dataframeemp_brut1[
        'QTD vendida Empresa I']
    dataframeemp_brut2['Preço'] = [a[2] for a in cart_m2_ont.acoes]
    dataframeemp_brut2['Ticker'] = [a[0] for a in cart_m2_ont.acoes]
    dataframeemp_brut2['QTD comprada Empresa II'] = [a[1] if a[6] in ['C', 'D'] else 0 for a in cart_m2_ont.acoes]
    dataframeemp_brut2['QTD vendida Empresa II'] = [-1 * a[1] if a[6] == 'V' else 0 for a in cart_m2_ont.acoes]
    dataframeemp_brut2['QTD bloq. Empresa II'] = [a[4] for a in cart_m2_ont.acoes]
    dataframeemp_brut2['QTD doada Empresa II'] = [a[4] if a[6] == 'D' else 0 for a in cart_m2_ont.acoes]
    dataframeemp_brut2['QTD tomada Empresa II'] = [a[4] if a[6] == 'T' else 0 for a in cart_m2_ont.acoes]
    dataframeemp_brut2['QTD vendida Empresa II'] = dataframeemp_brut2['QTD tomada Empresa II'] - dataframeemp_brut2['QTD vendida Empresa II']
    dataframeemp_brut2['QTD livre Empresa II'] = dataframeemp_brut2['QTD comprada Empresa II'] + dataframeemp_brut2[
        'QTD vendida Empresa II']
    dataframeemp1 = dataframeemp_brut1.groupby('Ticker').sum()
    precos = dataframeemp_brut2.groupby('Ticker').mean()
    dataframeemp2 = dataframeemp_brut2.groupby('Ticker').sum()
    dataframeemp2['Preço'] = precos['Preço']
    dataframeemp = dataframeemp1.merge(dataframeemp2, on=['Ticker'], how='left')
    dataframeemp = dataframeemp.fillna(0)
    dataframe_tot = dataframeemp.merge(dataframebtg, on=['Ticker'])
    dataframe_tot['Qtd a doar Empresa I'] = round(
        ((dataframe_tot['QTD livre Empresa I'] + dataframe_tot['QTD bloq. Empresa I']) * 0.7), 0) - dataframe_tot[
                                            'QTD bloq. Empresa I']
    dataframe_tot['Qtd a doar Empresa II'] = round(
        ((dataframe_tot['QTD livre Empresa II'] + dataframe_tot['QTD bloq. Empresa II']) * 0.7), 0) - \
                                         dataframe_tot['QTD bloq. Empresa II']

    broker = dataframe_tot[dataframe_tot['Taxa média'] >= 0.0010]
    lista_broker = [list(broker['Ticker']), list(broker['Qtd a doar Empresa I']), list(broker['Qtd a doar Empresa II'])]
    dataframebtcs = pd.DataFrame(btcs)
    dataframebtcs['Ticker'] = dataframebtcs[0]
    dataframebtcs['A Liquidar Empresa I'] = [int(a[7]) if a[2] == 'Empresa I' else 0 for a in btcs]
    dataframebtcs['A Liquidar Empresa II'] = [int(a[7]) if a[2] == 'Empresa II' else 0 for a in btcs]
    dataframebtcs = dataframebtcs.groupby('Ticker', as_index=False).sum()
    dataframe_tot = dataframe_tot.merge(dataframebtcs, on='Ticker', how='left')
    dataframe_tot = dataframe_tot.fillna(0)
    dataframe_tot['QTD livre Empresa I'] = dataframe_tot['QTD livre Empresa I'] + dataframe_tot['A Liquidar Empresa I']
    dataframe_tot['QTD livre Empresa II'] = dataframe_tot['QTD livre Empresa II'] + dataframe_tot['A Liquidar Empresa II']
    lista_tabela = [list(dataframe_tot['Ticker']), list(dataframe_tot['QTD livre Empresa I']), list(dataframe_tot['QTD livre Empresa II']),
                    list(dataframe_tot['QTD doada Empresa I']),
                    list(dataframe_tot['QTD doada Empresa II']), list(dataframe_tot['Taxa média']),
                    [round(i, 2) for i in list(dataframe_tot['SIR'])]]
    pl_alugadoi = sum([a * b for a, b in zip(list(dataframe_tot['QTD doada Empresa I']), list(dataframe_tot['Preço']))])/float(cart_f1_ont.pl) * 100
    pl_alugadoii = sum([a * b for a, b in zip(list(dataframe_tot['QTD doada Empresa II']), list(dataframe_tot['Preço']))])/float(cart_f2_ont.pl) * 100
    media_rem1 = sum([float(j[4].replace(',', '.')) / 100 for j in btcs if j[2] == 'Empresa I']) / len([j[4] for j in btcs if j[2] == 'Empresa I']) * 100
    media_rem2 = sum([float(j[4].replace(',', '.')) / 100 for j in btcs if j[2] == 'Empresa II']) / len([j[4] for j in btcs if j[2] == 'Empresa II']) * 100
    wb = openpyxl.load_workbook(
        rf'C:\Users\{caminho_Empresa}\ROTINAS\Relatório - BTC\Template - Relatório BTC.xlsx')
    ws = wb.sheetnames
    planilha = wb[ws[0]]
    s = 3
    for f in lista_tabela:
        x = 0
        for k in f:
            planilha.cell(row=5 + x, column=s).value = naonul(k)
            x += 1
        s += 1
    if btcs != []:
        v = 0
        b = 0
        for j in btcs:
            planilha.cell(row=5 + v, column=11).value = j[0]
            planilha.cell(row=5 + v, column=12).value = j[1]
            planilha.cell(row=5 + v, column=13).value = j[2]
            planilha.cell(row=5 + v, column=14).value = int(j[3])
            planilha.cell(row=5 + v, column=15).value = float(j[4].replace(',', '.')) / 100
            planilha.cell(row=5 + v, column=16).value = j[5]
            planilha.cell(row=5 + v, column=17).value = j[7]
            v += 1
            if j[5] in prox_5():
                planilha.cell(row=5 + b, column=20).value = j[0]
                planilha.cell(row=5 + b, column=21).value = j[2]
                planilha.cell(row=5 + b, column=22).value = int(j[3])
                planilha.cell(row=5 + b, column=23).value = j[5]
                b += 1
    bb = 0
    if gara != []:
        for k in gara:
            planilha.cell(row=17 + bb, column=20).value = k[0]
            planilha.cell(row=17 + bb, column=21).value = k[1]
            planilha.cell(row=17 + bb, column=22).value = k[2]
            planilha.cell(row=17 + bb, column=23).value = k[3]
            bb += 1
    h = 25
    for f in lista_broker:
        hh = 0
        for k in f:
            planilha.cell(row=5 + hh, column=h).value = naonul(k)
            hh += 1
        h += 1

    wb.save(
        rf'C:\Users\{caminho_Empresa}\ROTINAS\Relatório - BTC\Enviados\Relatório BTC - {data}.xlsx')
    outlook = win32.Dispatch('outlook.application')
    mail = outlook.CreateItem(0)
    mail.To = 'mail1'
    mail.CC = 'mail2'
    mail.Subject = f'Relatório BTC - {dia}/{mes}/{ano}'
    mail.Body = f'Bom dia! \n\nSegue anexa a planilha com o relatório de BTC relativo ao dia de hoje.\n\n O Empresa I está com {round(pl_alugadoi, 2)}% do PL alugado a uma taxa média de {round(media_rem1, 2)}%. ' \
                f'\n\nO Empresa II está com {round(pl_alugadoii, 2)}% do PL alugado a uma taxa média de {round(media_rem2, 2)}%.\n\nAtt,'
    mail.Attachments.Add(
        rf'C:\Users\{caminho_Empresa}\ROTINAS\Relatório - BTC\Enviados\Relatório BTC - {data}.xlsx')
    mail.Display(True)


def comando13():
    outlook = win32.Dispatch('outlook.application')
    mail = outlook.CreateItem(0)
    mail.To = 'mail2'
    mail.Bcc = "mail1"
    mail.Subject = f'Cotas Empresa - {ontem_bonito}'
    mail.Body = f"Prezados, bom dia! \n \n Segue as últimas cotas disponíveis dos fundos , referentes ao dia {ontem_bonito}: \n \n Empresa I FIC FIA: {cart_f1_ont.cota} \n \n  Empresa II FIC FIA: {cart_f2_ont.cota}\n \n Att,"
    mail.Display(True)
    try:
        ima = xlrd.open_workbook(
            fr'C:\Users\{caminho_Empresa}\ROTINAS\Balanco\IMA-B\{ontem}.xls')
    except:
        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
        time.sleep(2)
        driver.implicitly_wait(10)
        driver.maximize_window()
        driver.get("https://www.anbima.com.br/informacoes/ima/arqs/ima_completo.xls")
        time.sleep(5)
        os.rename(fr'C:\Users\{os.getlogin()}\Downloads\ima_completo.xls',
                fr'C:\Users\{caminho_Empresa}\ROTINAS\Balanco\IMA-B\{ontem}.xls')
        ima = xlrd.open_workbook(
            fr'C:\Users\{caminho_Empresa}\ROTINAS\Balanco\IMA-B\{ontem}.xls')
    tab1imab = ima.sheet_by_index(6)
    imab = tab1imab.cell_value(4, 19) / 100
    fat_imab = (1 + imab) ** (1 / 252)
    chromeOptions = webdriver.ChromeOptions()
    prefs = {'safebrowsing.enabled': 'false'}
    chromeOptions.add_experimental_option("prefs", prefs)
    time.sleep(2)
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
    driver.maximize_window()
    driver.implicitly_wait(10)
    driver.get(
        "https://www.b3.com.br/pt_br/market-data-e-indices/servicos-de-dados/market-data/consultas/mercado-de-derivativos/indicadores/indicadores-financeiros/")
    driver.switch_to.frame("bvmf_iframe")
    dolarrr = driver.find_element(By.XPATH,
                                  '//*[@id="divContainerIframeB3"]/form/div/div/div/div/div[2]/div[3]/div/div/h4').text
    driver.get("https://sistemaswebb3-listados.b3.com.br/indexStatisticsPage/variation/IBOVESPA?language=pt-br")
    ibovvv = driver.find_element(By.XPATH,
                                 '//*[@id="divContainerIframeB3"]/div/div[1]/form/div[1]/div/table/tbody/tr[1]/td[2]').text
    driver.get("https://sistemaswebb3-listados.b3.com.br/indexStatisticsPage/variation/IBXL?language=pt-br")
    ibxxx = driver.find_element(By.XPATH,'//*[@id="divContainerIframeB3"]/div/div[1]/form/div[1]/div/table/tbody/tr[1]/td[2]').text
    with open(
            fr'C:\Users\{caminho_Empresa}\ROTINAS\Relatorios Diarios\Bases de dados\IPCA_atual.txt',
            'r') as ip:
        ipc = ip.readline()
        ipcc = ipc.replace('IPCA Projetado Atual: ', '')
        ipca = float(ipcc.replace(',', '.'))
    with open(
            fr'C:\Users\{caminho_Empresa}\ROTINAS\Relatorios Diarios\Bases de dados\Arquivos CDI\{ontem}.txt',
            'r', encoding='latin-1') as cd:
        a = cd.readlines()
        c = a[2].split(";", 1000)
        fat_cdi = float(c[2].replace(',', '.'))
    cdi = fat_cdi - 1
    driver.quit()
    ibovv = ibovvv.replace('.', '')
    ibov = float(ibovv.replace(',', '.'))
    ibxx = ibxxx.replace('.', '')
    ibx = float(ibxx.replace(',', '.'))
    dolarr = dolarrr.replace(' (R$/US$)', '')
    dolar = float(dolarr.replace(',', '.'))
    fat_ipca = (1 + ipca) ** (1 / 22)
    bench = fat_ipca * fat_imab
    himab = pd.read_excel(
        rf'C:\Users\{caminho_Empresa}\ROTINAS\Relatorios Diarios\Bases de dados\Hist Fator IMA-B.xlsx',
        0)
    axhimab = pd.DataFrame({'Data': [yesterday], 'Cotacao': [imab]})
    himab = pd.concat([himab, axhimab])
    himab.to_excel(
        rf'C:\Users\{caminho_Empresa}\ROTINAS\Relatorios Diarios\Bases de dados\Hist Fator IMA-B.xlsx',
        index=False)
    hipca = pd.read_excel(
        rf'C:\Users\{caminho_Empresa}\ROTINAS\Relatorios Diarios\Bases de dados\Hist Fator IPCA.xlsx',
        0)
    axhipca = pd.DataFrame({'Data': [yesterday], 'Cotacao': [ipca]})
    hipca = pd.concat([hipca, axhipca])
    hipca.to_excel(
        rf'C:\Users\{caminho_Empresa}\ROTINAS\Relatorios Diarios\Bases de dados\Hist Fator IPCA.xlsx',
        index=False)
    hcdi = pd.read_excel(
        rf'C:\Users\{caminho_Empresa}\ROTINAS\Relatorios Diarios\Bases de dados\Hist Fator CDI.xlsx',
        0)
    axhcdi = pd.DataFrame({'Data': [yesterday], 'Cotacao': [cdi]})
    hcdi = pd.concat([hcdi, axhcdi])
    hcdi.to_excel(
        rf'C:\Users\{caminho_Empresa}\ROTINAS\Relatorios Diarios\Bases de dados\Hist Fator CDI.xlsx',
        index=False)
    h = pd.read_excel(
        rf'C:\Users\{caminho_Empresa}\ROTINAS\Relatorios Diarios\Bases de dados\Hist Ibov.xlsx',
        0)
    axh = pd.DataFrame({'Data': [yesterday], 'Cotacao': [ibov]})
    h = pd.concat([h, axh])
    h.to_excel(
        rf'C:\Users\{caminho_Empresa}\ROTINAS\Relatorios Diarios\Bases de dados\Hist Ibov.xlsx',
        index=False)
    x = pd.read_excel(
        rf'C:\Users\{caminho_Empresa}\ROTINAS\Relatorios Diarios\Bases de dados\Hist IBX.xlsx',
        0)
    listaibx = list(x['Cotacao'])
    var_ibx = ibx / listaibx[812] - 1
    axx = pd.DataFrame({'Data': [yesterday], 'Cotacao': [ibx]})
    x = pd.concat([x, axx])
    x.to_excel(
        rf'C:\Users\{caminho_Empresa}\ROTINAS\Relatorios Diarios\Bases de dados\Hist IBX.xlsx',
        index=False)
    i = pd.read_excel(
        rf'C:\Users\{caminho_Empresa}\ROTINAS\Relatorios Diarios\Bases de dados\Hist CDI.xlsx',
        0)
    listacdi = list(i['Cotacao'])
    cdi_novo = ((1 + listacdi[len(listacdi) - 1]) * (1 + cdi)) - 1
    axi = pd.DataFrame({'Data': [yesterday], 'Cotacao': [cdi_novo]})
    i = pd.concat([i, axi])
    i.to_excel(
        rf'C:\Users\{caminho_Empresa}\ROTINAS\Relatorios Diarios\Bases de dados\Hist CDI.xlsx',
        index=False)
    k = pd.read_excel(
        rf'C:\Users\{caminho_Empresa}\ROTINAS\Relatorios Diarios\Bases de dados\Hist Benchmark.xlsx',
        0)
    listabench = list(k['Cotacao'])
    bench_novo = ((1 + listabench[len(listabench) - 1]) * bench) - 1
    var_bench = bench_novo / listabench[0] - 1
    axk = pd.DataFrame({'Data': [yesterday], 'Cotacao': [bench_novo]})
    k = pd.concat([k, axk])
    k.to_excel(
        rf'C:\Users\{caminho_Empresa}\ROTINAS\Relatorios Diarios\Bases de dados\Hist Benchmark.xlsx',
        index=False)
    emp1 = pd.read_excel(
        rf'C:\Users\{caminho_Empresa}\ROTINAS\Relatorios Diarios\Bases de dados\Hist Empresa I.xlsx',
        0)
    emp1_nov = float(cart_f1_ont.cota) - 1
    axempI = pd.DataFrame(
        {'Data': [yesterday], 'Fundo': [emp1_nov], 'Benchmark': [var_bench], 'Ibovespa': [ibov], 'CDI': [cdi_novo]})
    emp1 = pd.concat([emp1, axempI])
    emp1.to_excel(
        rf'C:\Users\{caminho_Empresa}\ROTINAS\Relatorios Diarios\Bases de dados\Hist Empresa I.xlsx',
        index=False)

    emp1i = pd.read_excel(
        rf'C:\Users\{caminho_Empresa}\ROTINAS\Relatorios Diarios\Bases de dados\Hist Empresa II.xlsx',
        0)
    emp1i_nov = float(cart_f2_ont.cota) - 1
    axempIi = pd.DataFrame({'Data': [yesterday], 'Empresa FIC FIA': [emp1i_nov], 'Benchmark': [var_ibx],
                            'Ibov_Desde_Inicio': [ibov], 'CDI_Desde_Inicio': [cdi_novo]})
    emp1i = pd.concat([emp1i, axempIi])
    emp1i.to_excel(
        rf'C:\Users\{caminho_Empresa}\ROTINAS\Relatorios Diarios\Bases de dados\Hist Empresa II.xlsx',
        index=False)
    # atualiza as carteiras
    cart_emp1 = cart_m1_ont.pega_carteira(dolar)
    cart_emp2 = cart_m2_ont.pega_carteira(dolar)
    lista_ativos_emp1 = [list(cart_emp1[a]) for a in ['Ticker', 'Quantidade', 'Preco','Financeiro', '% do PL', 'Cot dolar', '% Aluguel', '% Bloq']]
    lista_ativos_emp2 = [list(cart_emp2[a]) for a in ['Ticker', 'Quantidade', 'Preco','Financeiro', '% do PL', 'Cot dolar', '% Aluguel', '% Bloq']]
    # atualizacao de becnhmarks
    empI = pd.read_excel(
        rf'C:\Users\{caminho_Empresa}\ROTINAS\Relatorios Diarios\Bases de dados\Hist Empresa I.xlsx',
        0)
    ano_mes1 = [f'{str(a)[5:7]}/{str(a)[0:4]}' for a in list(empI['Data'])]
    empI['Ref'] = ano_mes1
    empI['Fundo'] = empI['Fundo'] + 1
    empI['Benchmark'] = empI['Benchmark'] + 1
    rent_mensal1 = pd.DataFrame(columns=['Ref', 'Fundo', 'Benchmark', 'Ibovespa', 'CDI'])
    rent_mensal1['Ref'] = ['Início']
    rent_mensal1['Fundo'] = [1]
    for i in set(empI['Ref']):
        df = empI[empI['Ref'] == i]
        ponta = df.tail(n=1)
        rent_mensal1 = pd.concat([rent_mensal1, ponta], join='inner')
    rent_mensal1 = rent_mensal1.sort_index(ascending=True)
    lista_rents1 = [list(rent_mensal1['Ref']), list(rent_mensal1['Fundo']), list(rent_mensal1['Benchmark']),
                    list(rent_mensal1['Ibovespa']), list(rent_mensal1['CDI'])]
    empII = pd.read_excel(
        rf'C:\Users\{caminho_Empresa}\ROTINAS\Relatorios Diarios\Bases de dados\Hist Empresa II.xlsx',
        0)
    ano_mes2 = [f'{str(a)[5:7]}/{str(a)[0:4]}' for a in list(empII['Data'])]
    empII['Ref'] = ano_mes2
    empII['Empresa FIC FIA'] = empII['Empresa FIC FIA'] + 1
    empII['Benchmark'] = empII['Benchmark'] + 1
    rent_mensal2 = pd.DataFrame(
        columns=['Ref', 'Empresa FIC FIA', 'Benchmark', 'Ibov_Desde_Inicio', 'CDI_Desde_Inicio'])
    rent_mensal2['Ref'] = ['Início']
    rent_mensal2['Empresa FIC FIA'] = [1]
    for i in set(empII['Ref']):
        df = empII[empII['Ref'] == i]
        ponta = df.tail(n=1)
        rent_mensal2 = pd.concat([rent_mensal2, ponta], join='inner')
    rent_mensal2 = rent_mensal2.sort_index(ascending=True)
    lista_rents2 = [list(rent_mensal2['Ref']), list(rent_mensal2['Empresa FIC FIA']), list(rent_mensal2['Benchmark']),
                    list(rent_mensal2['Ibov_Desde_Inicio']), list(rent_mensal2['CDI_Desde_Inicio'])]
    #  Atualização de arquivos na planilha

    wb = openpyxl.load_workbook(
        rf'C:\Users\{caminho_Empresa}\ROTINAS\Relatorios Diarios\Planilha base - Relatório Diário.xlsx')
    ws = wb.sheetnames
    planilha = wb[ws[0]]
    planilha2 = wb[ws[1]]
    opcoes_empI = [[a[0], a[4]] for a in cart_m1_ont.opcoes]
    opcoes_empIi = [[a[0], a[4]] for a in cart_m2_ont.opcoes]
    cvm1 = sum([(a[1] + a[4]) * a[2] for a in cart_m1_ont.acoes if a[6] in ['D', 'C']]) / float(
        cart_m1_ont.pl)
    gross1 = sum([(a[1] + a[4]) * a[2] if a[6] in ['D', 'C', 'T'] else float(a[3]) * -1 for a in
                  cart_m1_ont.acoes]) / float(
        cart_f1_ont.pl)
    net1 = sum([(a[1] + a[4]) * a[2] if a[6] in ['D', 'C', 'V'] else -1 * (a[1] + a[4]) * a[2] for a in
                cart_m1_ont.acoes]) / float(
        cart_f1_ont.pl)
    cvm2 = sum([(a[1] + a[4]) * a[2] for a in cart_m2_ont.acoes if a[6] in ['D', 'C']]) / float(
        cart_m2_ont.pl)
    gross2 = sum([(a[1] + a[4]) * a[2] if a[6] in ['D', 'C', 'T'] else float(a[3]) * -1 for a in
                  cart_m2_ont.acoes]) / float(
        cart_f2_ont.pl)
    net2 = sum([(a[1] + a[4]) * a[2] if a[6] in ['D', 'C', 'V'] else -1 * (a[1] + a[4]) * a[2] for a in
                cart_m2_ont.acoes]) / float(
        cart_f2_ont.pl)
    cx = cart_m1_ont.caixa
    cxf = cart_f1_ont.caixa
    planilha.cell(row=7, column=11).value = float(cart_f1_ont.cota)
    planilha.cell(row=8, column=11).value = float(cart_f1_ont.pl)
    planilha.cell(row=9, column=11).value = (float(cart_f1_ont.cota) / float(cart_f1_antont.cota)) - 1
    planilha.cell(row=43, column=7).value = cx[0]
    planilha.cell(row=44, column=7).value = cxf[0]
    planilha.cell(row=45, column=7).value = cx[1] / dolar
    planilha.cell(row=46, column=7).value = cx[1]
    planilha.cell(row=47, column=7).value = sum([a[2] for a in cart_m1_ont.titpublico]) + sum([a[2] for a in cart_f1_ont.titpublico])
    planilha.cell(row=43, column=4).value = cvm1
    planilha.cell(row=44, column=4).value = gross1
    planilha.cell(row=45, column=4).value = net1
    s = 4
    for f in lista_ativos_emp1:
        x = 0
        for k in f:
            planilha.cell(row=13 + x, column=s).value = k
            x += 1
        s += 1
    zz = 10
    for f in lista_rents1:
        z = 0
        for k in f:
            planilha.cell(row=zz, column=18 + z).value = k
            z += 1
        zz += 1
    ss = 4
    for f in lista_ativos_emp2:
        xx = 0
        for k in f:
            planilha2.cell(row=13 + xx, column=ss).value = k
            xx += 1
        ss += 1
    cx2 = cart_m2_ont.caixa
    cx2f = cart_f2_ont.caixa
    planilha2.cell(row=7, column=11).value = float(cart_f2_ont.cota)
    planilha2.cell(row=8, column=11).value = float(cart_f2_ont.pl)
    planilha2.cell(row=9, column=11).value = (float(cart_f2_ont.cota) / float(cart_f2_antont.cota)) - 1
    planilha2.cell(row=43, column=7).value = cx2[0]
    planilha2.cell(row=44, column=7).value = cx2f[0]
    planilha2.cell(row=45, column=7).value = cx2[1] / dolar
    planilha2.cell(row=46, column=7).value = cx2[1]
    planilha2.cell(row=47, column=7).value = sum([a[2] for a in cart_m2_ont.titpublico]) + sum([a[2] for a in cart_f2_ont.titpublico])
    planilha2.cell(row=43, column=4).value = cvm2
    planilha2.cell(row=44, column=4).value = gross2
    planilha2.cell(row=45, column=4).value = net2
    op = 0
    for x in opcoes_empI:
        planilha.cell(row=43 + op, column=10).value = x[0]
        planilha.cell(row=43 + op, column=11).value = x[1]
        op += 1

    opp = 0
    for h in opcoes_empIi:
        planilha2.cell(row=43 + opp, column=10).value = h[0]
        planilha2.cell(row=43 + opp, column=11).value = h[1]
        opp += 1
    zc = 10
    for f in lista_rents2:
        c = 0
        for k in f:
            planilha2.cell(row=zc, column=18 + c).value = k
            c += 1
        zc += 1
    wb.save(
        rf'C:\Users\{caminho_Empresa}\ROTINAS\Relatorios Diarios\Excel\{ontem}.xlsx')
    time.sleep(3)
    excel = win32.Dispatch('Excel.Application')
    wb = excel.Workbooks.Open(
        rf'C:\Users\{caminho_Empresa}\ROTINAS\Relatorios Diarios\Excel\{ontem}.xlsx')
    sheet = wb.Sheets[0]
    excel.visible = 1
    copyrange = sheet.Range('A1:Q99')
    copyrange.CopyPicture(Appearance=1, Format=2)
    ImageGrab.grabclipboard().save(
        rf'C:\Users\{caminho_Empresa}\ROTINAS\Relatorios Diarios\Empresa I\Relatório diário Empresa I - {ontem_arq}.pdf')
    sheet2 = wb.Sheets[1]
    copyrange2 = sheet2.Range('A1:Q99')
    copyrange2.CopyPicture(Appearance=1, Format=2)
    ImageGrab.grabclipboard().save(
        rf'C:\Users\{caminho_Empresa}\ROTINAS\Relatorios Diarios\Empresa II\Relatório diário Empresa II - {ontem_arq}.pdf')
    excel.Quit()
    outlook = win32.Dispatch('outlook.application')
    mail = outlook.CreateItem(0)
    mail.To = 'mail1'
    mail.CC = 'mail2'
    mail.Subject = f'Relatórios Diários - {dia}/{mes}/{ano}'
    mail.Body = f'Bom dia!\n\nSeguem os relatórios diários referentes ao dia {ontem_bonito}.\n\nAtt,'
    mail.Attachments.Add(
        rf'C:\Users\{caminho_Empresa}\ROTINAS\Relatorios Diarios\Empresa I\Relatório diário Empresa I - {ontem_arq}.pdf')
    mail.Attachments.Add(
        rf'C:\Users\{caminho_Empresa}\ROTINAS\Relatorios Diarios\Empresa II\Relatório diário Empresa II - {ontem_arq}.pdf')
    #mail.Display()
    #mail = outlook.CreateItem(0)
    #mail.CC = 'mail2'
    #mail.Body = f'Bom dia!\n\nSegue posição para Anijes Empreendimentos referente ao dia {ontem_bonito}.\n\n\nAtt,'
    mail.Display()


def comando14():
    dados14 = []
    prompt14 = tk.Tk()
    c14 = tk.Canvas(prompt14, width=300, height=225, relief='raised', bg=preto_Empresa)
    c14.pack()
    v141 = tk.StringVar(prompt14)
    l10 = tk.Label(prompt14, text='Auxiliar - Taxa ADM e Rebates', font=('helvetica', 14), bg=preto_Empresa, fg=amarelo_Empresa)
    c14.create_window(150, 25, window=l10)
    l141 = tk.Label(prompt14, text="Insira o mês da planilha auxiliar: Ex (AAAAMM)",
                    font=('helvetica', 10), bg=preto_Empresa, fg=amarelo_Empresa)
    c14.create_window(150, 70, window=l141)
    ent141 = tk.Entry(prompt14, textvariable=v141)
    c14.create_window(150, 95, window=ent141)

    def pegainput14():
        global w141
        w141 = v141.get()
        dados14.append(w141)
        prompt14.destroy()
        root1.quit()

    bot14 = tk.Button(prompt14, text='Gerar >', command=pegainput14, bg=branco_Empresa, fg=preto_Empresa,
                      font=('helvetica', 9, 'bold'))
    c14.create_window(150, 200, window=bot14)
    prompt14.mainloop()
    tk.messagebox.showinfo('Espere', 'Esse processo pode demorar. Por favor, aguarde.')
    mes = int(dados14[0])
    pls = []
    cotas = []
    for i in range(int(f'{mes - 1}27'), int(f'{mes}32')):
        try:
            k = str(i)
            masteri = Carteira(i, 'm1')
            fici = Carteira(i, 'f1')
            masterii = Carteira(i, 'm2')
            ficii = Carteira(i, 'f2')
            if cart_f1_ont.pl != None:
                pls.append([f'{k[6]}{k[7]}/{k[4]}{k[5]}/{k[0]}{k[1]}{k[2]}{k[3]}', masteri.pl, fici.pl,
                            masterii.pl, ficii.pl])
                cotas.append([f'{k[6]}{k[7]}/{k[4]}{k[5]}/{k[0]}{k[1]}{k[2]}{k[3]}', masteri.cota, fici.cota,
                            masterii.cota, ficii.cota])
        except:
            pass
    wb = openpyxl.load_workbook(
        rf'C:\Users\{caminho_Empresa}\ROTINAS\Auxiliar - Taxa ADM\Template - Aux. Taxa ADM.xlsx')
    ws = wb.sheetnames
    planilha = wb[ws[0]]
    s = 5
    for f in pls:
        planilha.cell(row=s, column=3).value = f[0]
        planilha.cell(row=s, column=4).value = f[1]
        planilha.cell(row=s, column=5).value = f[2]
        planilha.cell(row=s, column=6).value = f[3]
        planilha.cell(row=s, column=7).value = f[4]
        s += 1
    wb.save(
        rf'C:\Users\{caminho_Empresa}\ROTINAS\Auxiliar - Taxa ADM\{dados14[0]}.xlsx')
    wbc = openpyxl.load_workbook(
        rf'C:\Users\{caminho_Empresa}\ROTINAS\Rebates\Template - Cotas.xlsx')
    wsc = wbc.sheetnames
    planilhac = wbc[wsc[0]]
    s = 2
    fundos = {1: 652091, 2:387894, 3:454168, 4:454109}
    for f in range(1, 5):
        for i, k in zip([a[0] for a in cotas], [a[f] for a in cotas]):
            planilhac.cell(row=s, column=1).value = fundos[f]
            planilhac.cell(row=s, column=2).value = i
            planilhac.cell(row=s, column=3).value = float(k)
            planilhac.cell(row=s, column=4).value = 0
            planilhac.cell(row=s, column=5).value = 0
            s += 1
    wbc.save(
        rf'C:\Users\{caminho_Empresa}\ROTINAS\Rebates\Histórico de cotas para importação\{dados14[0]}.xlsx')
    tk.messagebox.showinfo('Auxiliar - Taxa ADM e Rebates',
                           'A planilha foi criada e enconra-se na Pasta "Auxiliar - Taxa ADM", com o título de seu mês')


def comando15():
    dados15 = []
    prompt15 = tk.Tk()
    c15 = tk.Canvas(prompt15, width=300, height=225, relief='raised', bg=preto_Empresa)
    c15.pack()
    v151 = tk.StringVar(prompt15)
    l15 = tk.Label(prompt15, text='Auxiliar - Taxa ADM', font=('helvetica', 14), bg=preto_Empresa, fg=amarelo_Empresa)
    c15.create_window(150, 25, window=l15)
    l151 = tk.Label(prompt15, text="Insira o último dia útil do Mês: Ex (AAAAMMDD)",
                    font=('helvetica', 10), bg=preto_Empresa, fg=amarelo_Empresa)
    c15.create_window(150, 70, window=l151)
    ent151 = tk.Entry(prompt15, textvariable=v151)
    c15.create_window(150, 95, window=ent151)

    def pegainput15():
        global w151
        w151 = v151.get()
        dados15.append(w151)
        prompt15.destroy()
        root1.quit()

    bot15 = tk.Button(prompt15, text='Gerar >', command=pegainput15, bg=branco_Empresa, fg=preto_Empresa,
                      font=('helvetica', 9, 'bold'))
    c15.create_window(150, 200, window=bot15)
    prompt15.mainloop()
    mes = dados15[0]
    with open(
            fr'C:\Users\{caminho_Empresa}\ROTINAS\Balanco\Carteiras XML\Empresa II MASTER FIA\{mes}.xml',
            'r') as m2:
        a = m2.readlines()
    with open(
            fr'C:\Users\{caminho_Empresa}\ROTINAS\Balanco\Carteiras XML\Empresa II FIC FIA\{mes}.xml',
            'r') as f2:
        b = f2.readlines()
    with open(
            fr'C:\Users\{caminho_Empresa}\ROTINAS\Auxiliar\XMLs mensais\XML Mensal - {mes} - Empresa II FIC FIA.xml',
            'w') as aa:
        aa.writelines(b)
    with open(
            fr'C:\Users\{caminho_Empresa}\ROTINAS\Auxiliar\XMLs mensais\XML Mensal - {mes} - Empresa II MASTER FIA.xml',
            'w') as bb:
        bb.writelines(a)
    outlook = win32.Dispatch('outlook.application')
    mail = outlook.CreateItem(0)
    mail.To = 'backoffice@mag.com.br'
    mail.Cc = 'mail2'
    mail.Attachments.Add(
        fr'C:\Users\{caminho_Empresa}\ROTINAS\Auxiliar\XMLs mensais\XML Mensal - {mes} - Empresa II MASTER FIA.xml')
    mail.Attachments.Add(
        fr'C:\Users\{caminho_Empresa}\ROTINAS\Auxiliar\XMLs mensais\XML Mensal - {mes} - Empresa II FIC FIA.xml')
    mail.Subject = f'Carteiras XML Empresa - {mes[4:6]}/{mes[0:4]}'
    mail.Body = f"Prezados, bom dia! \n \nSeguem as últimas carteiras dos fundos da Empresa, referentes ao último mês.  \n \nAtt,"
    mail.Display()


def comando16():
    wb = openpyxl.load_workbook(
        rf'C:\Users\{caminho_Empresa}\ROTINAS\Relatório de Risco\Template - Risco.xlsx')
    ws = wb.sheetnames
    planilha = wb[ws[0]]
    planilha2 = wb[ws[1]]
    planilha3 = wb[ws[2]]
    ativos_emp1 = cart_m1_ont.pega_carteira(5)
    lista_ativos_emp1 = [list(ativos_emp1[a]) for a in
                         ['Ticker', 'Quantidade', 'Preco', 'Financeiro', '% do PL', 'Cot dolar', '% Aluguel', '% Bloq']]
    ativos_emp2 = cart_m2_ont.pega_carteira(5)
    lista_ativos_emp2 = [list(ativos_emp2[a]) for a in
                         ['Ticker', 'Quantidade', 'Preco', 'Financeiro', '% do PL', 'Cot dolar', '% Aluguel', '% Bloq']]
    setores = pd.read_excel(
        fr'C:\Users\{caminho_Empresa}\ROTINAS\Relatório de Risco\Setores com ticker numerado.xlsx',
        0)
    setorial1 = ativos_emp1.merge(setores, on='Ticker', how='left')
    setorial1 = setorial1.groupby(['SETOR ECONÔMICO'], as_index=False).sum()
    setorial1 = setorial1.drop(columns=[a for a in list(setorial1) if a not in ['SETOR ECONÔMICO', '% do PL']])
    setorial1 = setorial1.rename(columns={'% do PL': '% do PL 1'})
    setorial2 = ativos_emp2.merge(setores, on='Ticker', how='left')
    setorial2 = setorial2.groupby(['SETOR ECONÔMICO'], as_index=False).sum()
    setorial2 = setorial2.drop(columns=[a for a in list(setorial2) if a not in ['SETOR ECONÔMICO', '% do PL']])
    setorial2 = setorial2.rename(columns={'% do PL': '% do PL 2'})
    setorial = setorial1.merge(setorial2, on='SETOR ECONÔMICO', how='left')
    setorial = setorial.sort_values('% do PL 1', ascending=False)
    s = 4
    for f in lista_ativos_emp1:
        x = 0
        for k in f:
            planilha.cell(row=13 + x, column=s).value = k
            x += 1
        s += 1
    ss = 4
    for f in lista_ativos_emp2:
        xx = 0
        for k in f:
            planilha2.cell(row=13 + xx, column=ss).value = k
            xx += 1
        ss += 1
    r = 0
    for k in list(setorial['SETOR ECONÔMICO']):
        planilha3.cell(row=10 + r, column=11).value = k
        r += 1
    rr = 0
    for k in list(setorial['% do PL 1']):
        planilha3.cell(row=10 + rr, column=13).value = k
        rr += 1
    rrr = 0
    for k in list(setorial['% do PL 2']):
        planilha3.cell(row=10 + rrr, column=14).value = k
        rrr += 1
    planilha3.cell(row=7, column=13).value = (cart_m1_ont.caixa[0] + cart_f1_ont.caixa[0] +
                                              sum([a[2] for a in cart_m1_ont.titpublico]) +
                                              sum([a[2] for a in cart_f1_ont.titpublico])) / float(cart_f1_ont.pl)
    planilha3.cell(row=7, column=14).value = (cart_m2_ont.caixa[0] + cart_f2_ont.caixa[0] +
                                              sum([a[2] for a in cart_m2_ont.titpublico]) +
                                              sum([a[2] for a in cart_f2_ont.titpublico])) / float(cart_f2_ont.pl)
    planilha3.cell(row=8, column=13).value = cart_m1_ont.caixa[1] / float(cart_f1_ont.pl)
    planilha3.cell(row=8, column=14).value = cart_m2_ont.caixa[1] / float(cart_f2_ont.pl)
    wb.save(
        fr'C:\Users\{caminho_Empresa}\ROTINAS\Relatório de Risco\Histórico\{ontem}.xlsx')
    excel = win32.Dispatch('Excel.Application')
    wbx = excel.Workbooks.Open(
        fr'C:\Users\{caminho_Empresa}\ROTINAS\Relatório de Risco\Histórico\{ontem}.xlsx',
        UpdateLinks=3)
    sheet = wbx.Sheets[2]
    excel.visible = 1
    copyrange = sheet.Range('A1:O57')
    copyrange.CopyPicture(Appearance=1, Format=2)
    ImageGrab.grabclipboard().save(
        fr'C:\Users\{caminho_Empresa}\ROTINAS\Relatório de Risco\Histórico\Relatório de Risco - {ontem}.pdf',
        quality=100)
    excel.Quit()
    outlook = win32.Dispatch('outlook.application')
    mail = outlook.CreateItem(0)
    mail.To = 'mail1'
    mail.Subject = f'Relatório de Risco {dia}/{mes}/{ano}'
    mail.Cc = 'mail2'
    mail.Attachments.Add(
        fr'C:\Users\{caminho_Empresa}\ROTINAS\Relatório de Risco\Histórico\Relatório de Risco - {ontem}.pdf')
    mail.Body = f"Prezados, bom dia! \n\nSegue o relatório interno de risco da semana passada.\n\nAtt,"
    mail.Display()


def comando17():
    with open(
            fr'C:\Users\{caminho_Empresa}\ROTINAS\Relatorios Diarios\Bases de dados\Arquivos CDI\{ontem}.txt',
            'r', encoding='latin-1') as cd:
        a = cd.readlines()[2].split(";", 1000)
        rentabilidade = float(a[2].replace(',', '.'))
    carteira1_ont = cart_m1_ont.pega_carteira(5)
    carteira1_antont = cart_m1_antont.pega_carteira(5)
    carteira2_ont = cart_m2_ont.pega_carteira(5)
    carteira2_antont = cart_m2_antont.pega_carteira(5)
    lista_ativos_emp1_o = [[a.replace(' US', '') for a in list(carteira1_ont['Ticker'])],
                           list(carteira1_ont['Quantidade']), list(carteira1_ont['Preco']),
                           list(carteira1_ont['Financeiro']), list(carteira1_ont['% do PL'])]
    lista_ativos_emp1_a = [[a.replace(' US', '') for a in list(carteira1_antont['Ticker'])],
                           list(carteira1_antont['Quantidade']), list(carteira1_antont['Preco']),
                           list(carteira1_antont['Financeiro']), list(carteira1_antont['% do PL'])]
    lista_ativos_emp2_o = [[a.replace(' US', '') for a in list(carteira2_ont['Ticker'])],
                           list(carteira2_ont['Quantidade']), list(carteira2_ont['Preco']),
                           list(carteira2_ont['Financeiro']), list(carteira2_ont['% do PL'])]
    lista_ativos_emp2_a = [[a.replace(' US', '') for a in list(carteira2_antont['Ticker'])],
                           list(carteira2_antont['Quantidade']), list(carteira2_antont['Preco']),
                           list(carteira2_antont['Financeiro']), list(carteira2_antont['% do PL'])]
    carcap_empI_o = pd.DataFrame()
    carcap_empIi_o = pd.DataFrame()
    carcap_empI_a = pd.DataFrame()
    carcap_empIi_a = pd.DataFrame()
    for i, j in zip([cart_m1_ont, cart_f1_ont, cart_m2_ont, cart_f2_ont],
                    [cart_m1_antont, cart_f1_antont, cart_m2_antont, cart_f2_antont]):
        ont = pd.DataFrame(a for a in i.provisoes if
                           a[0] not in ['Empréstimo de Ações', 'Subscrições', 'Ações ou Opções',
                                        'Aplicação a Converter'])
        ont.insert(0, 'Data', muda_data(str(ontem)))
        ont[2] = [muda_data(a) for a in list(ont[2])]
        ont.insert(0, 'Fundo', i.nome)
        antont = pd.DataFrame(a for a in j.provisoes if
                              a[0] not in ['Empréstimo de Ações', 'Subscrições', 'Ações ou Opções',
                                           'Aplicação a Converter'])
        antont.insert(0, 'Data', muda_data(str(anteontem)))
        antont.insert(0, 'Fundo', j.nome)
        antont[2] = [muda_data(a) for a in list(antont[2])]
        if i in [cart_m1_ont, cart_f1_ont]:
            carcap_empI_o = pd.concat([carcap_empI_o, ont])
            carcap_empI_a = pd.concat([carcap_empI_a, antont])
        else:
            carcap_empIi_o = pd.concat([carcap_empIi_o, ont])
            carcap_empIi_a = pd.concat([carcap_empIi_a, antont])
    carcap_emp1_o = [list(carcap_empI_o[a]) for a in list(carcap_empI_o) if a != 1]
    carcap_emp1_a = [list(carcap_empI_a[a]) for a in list(carcap_empI_a) if a != 1]
    carcap_emp2_o = [list(carcap_empIi_o[a]) for a in list(carcap_empIi_o) if a != 1]
    carcap_emp2_a = [list(carcap_empIi_a[a]) for a in list(carcap_empIi_a) if a != 1]
    wb = openpyxl.load_workbook(
        rf'C:\Users\{caminho_Empresa}\ROTINAS\Batimento - PnL\Template - Batimento PnL.xlsx')
    ws = wb.sheetnames
    planilha = wb[ws[0]]
    planilha2 = wb[ws[1]]
    planilha3 = wb[ws[2]]
    c_1mo = cart_m1_ont.caixa
    c_1ma = cart_m1_antont.caixa
    c_1fo = cart_f1_ont.caixa
    c_1fa = cart_f1_antont.caixa
    c_2mo = cart_m2_ont.caixa
    c_2ma = cart_m2_antont.caixa
    c_2fo = cart_f2_ont.caixa
    c_2fa = cart_f2_antont.caixa
    # primeiro os valores de caixa e pl
    planilha.cell(row=17, column=3).value = c_1mo[0] + c_1fo[0] + sum([a[2] for a in cart_m1_ont.titpublico]) + sum([a[2] for a in cart_f1_ont.titpublico])
    planilha.cell(row=18, column=3).value = c_1mo[1]
    planilha.cell(row=17, column=9).value = c_1ma[0] + c_1fa[0] + sum([a[2] for a in cart_m1_antont.titpublico])+ sum([a[2] for a in cart_m1_antont.titpublico])
    planilha.cell(row=18, column=9).value = c_1ma[1]
    planilha2.cell(row=17, column=3).value = c_2mo[0] + c_2fo[0] + sum([a[2] for a in cart_m2_ont.titpublico]) + sum([a[2] for a in cart_f2_ont.titpublico])
    planilha2.cell(row=18, column=3).value = c_2mo[1]
    planilha2.cell(row=17, column=9).value = c_2ma[0] + c_2fa[0] + sum([a[2] for a in cart_m2_antont.titpublico]) + sum([a[2] for a in cart_f2_antont.titpublico])
    planilha2.cell(row=18, column=9).value = c_2ma[1]
    planilha.cell(row=13, column=3).value = float(cart_f1_ont.pl)
    planilha.cell(row=13, column=9).value = float(cart_f1_antont.pl)
    planilha2.cell(row=13, column=3).value = float(cart_f2_ont.pl)
    planilha2.cell(row=13, column=9).value = float(cart_f2_antont.pl)
    planilha.cell(row=12, column=3).value = float(cart_f1_ont.cotas_a_emitir)
    planilha.cell(row=12, column=9).value = float(cart_f1_antont.cotas_a_emitir)
    planilha2.cell(row=12, column=3).value = float(cart_f2_ont.cotas_a_emitir)
    planilha2.cell(row=12, column=9).value = float(cart_f2_antont.cotas_a_emitir)
    # agora os valores de renda variável:
    s = 2
    for f in lista_ativos_emp1_o:
        x = 0
        for k in f:
            planilha.cell(row=22 + x, column=s).value = k
            x += 1
        s += 1
    ss = 8
    for f in lista_ativos_emp1_a:
        xx = 0
        for k in f:
            planilha.cell(row=22 + xx, column=ss).value = k
            xx += 1
        ss += 1
    s = 2
    for f in lista_ativos_emp2_o:
        x = 0
        for k in f:
            planilha2.cell(row=22 + x, column=s).value = k
            x += 1
        s += 1
    ss = 8
    for f in lista_ativos_emp2_a:
        xx = 0
        for k in f:
            planilha2.cell(row=22 + xx, column=ss).value = k
            xx += 1
        ss += 1
    s = 2
    for f in lista_ativos_emp1_o:
        x = 0
        for k in f:
            planilha.cell(row=22 + x, column=s).value = k
            x += 1
        s += 1
    ss = 8
    for f in lista_ativos_emp1_a:
        xx = 0
        for k in f:
            planilha.cell(row=22 + xx, column=ss).value = k
            xx += 1
        ss += 1
    s = 2
    for f in lista_ativos_emp2_o:
        x = 0
        for k in f:
            planilha2.cell(row=22 + x, column=s).value = k
            x += 1
        s += 1
    ss = 8
    for f in lista_ativos_emp2_a:
        xx = 0
        for k in f:
            planilha2.cell(row=22 + xx, column=ss).value = k
            xx += 1
        ss += 1
    planilha.cell(row=51, column=5).value = sum([a[3] for a in cart_m1_ont.provisoes if
                                                 a[0] in ['Ações ou Opções', 'Valor Bovespa', 'Valor Repasse Bovespa']])
    planilha.cell(row=51, column=11).value = sum([a[3] for a in cart_m1_antont.provisoes if
                                                  a[0] in ['Ações ou Opções', 'Valor Bovespa',
                                                           'Valor Repasse Bovespa']])
    planilha2.cell(row=51, column=5).value = sum([a[3] for a in cart_m2_ont.provisoes if
                                                  a[0] in ['Ações ou Opções', 'Valor Bovespa',
                                                           'Valor Repasse Bovespa']])
    planilha2.cell(row=51, column=11).value = sum([a[3] for a in cart_m2_antont.provisoes if
                                                   a[0] in ['Ações ou Opções', 'Valor Bovespa',
                                                            'Valor Repasse Bovespa']])
    s = 2   # precisaríamos colocar também as oopções de d-1, pra que a planilha esteja 100% completa.
    for j in [[a[0] for a in cart_m1_ont.opcoes], [a[2] for a in cart_m1_ont.opcoes],
              [a[3] for a in cart_m1_ont.opcoes]]:
        c = 0
        for k in j:
            planilha.cell(row=112 + c, column=s).value = k
            c += 1
        s += 1
    s = 2
    for j in [[a[0] for a in cart_m2_ont.opcoes], [a[2] for a in cart_m2_ont.opcoes],
              [a[3] for a in cart_m2_ont.opcoes]]:
        c = 0
        for k in j:
            planilha2.cell(row=112 + c, column=s).value = k
            c += 1
        s += 1
    # variações no carcap
    s = 2
    for f in carcap_emp1_o:
        x = 0
        for k in f:
            planilha.cell(row=56 + x, column=s).value = k
            x += 1
        s += 1
    ss = 8
    for f in carcap_emp1_a:
        xx = 0
        for k in f:
            planilha.cell(row=56 + xx, column=ss).value = k
            xx += 1
        ss += 1
    s = 2
    for f in carcap_emp2_o:
        x = 0
        for k in f:
            planilha2.cell(row=56 + x, column=s).value = k
            x += 1
        s += 1
    ss = 8
    for f in carcap_emp2_a:
        xx = 0
        for k in f:
            planilha2.cell(row=56 + xx, column=ss).value = k
            xx += 1
        ss += 1
    planilha3.cell(row=6, column=4).value = ((cart_m1_antont.caixa[0] + cart_f1_antont.caixa[0]) * rentabilidade)/float(cart_f1_ont.pl)
    planilha3.cell(row=6, column=7).value = ((cart_m2_antont.caixa[0] + cart_f2_antont.caixa[0]) * rentabilidade)/float(cart_f2_ont.pl)
    planilha3.cell(row=2, column=4).value = muda_data(ontem)
    wb.save(
        rf'C:\Users\{caminho_Empresa}\ROTINAS\Batimento - PnL\Hist_novo\{muda_data(ontem).replace("/", "-")}.xlsx')
    excel = win32.Dispatch('Excel.Application')
    excel.Workbooks.Open(
        rf'C:\Users\{caminho_Empresa}\ROTINAS\Batimento - PnL\Hist_novo\{muda_data(ontem).replace("/", "-")}.xlsx',
        UpdateLinks=3)
    excel.visible = 1


def comando18():
    def teste(a):
        if a == 0:
            return '-'
        else:
            return f'{a}%'
    plsi = []
    plsii = []
    for k in range(int(ontem) - 10000, int(ontem)):
        try:
            fici = Carteira(k, 'f1')
            ficii = Carteira(k, 'f2')
            if float(fici.pl) >= 2:
                plsi.append(float(fici.pl))
                plsii.append(float(ficii.pl))
        except:
            pass
    excel = openpyxl.load_workbook(
        rf'C:\Users\{caminho_Empresa}\ROTINAS\Relatorios Diarios\Excel\{ontem}.xlsx',
        data_only=True)
    fundos = excel.sheetnames
    empItext = excel[fundos[0]]
    empIitext = excel[fundos[1]]
    with open(
            rf'C:\Users\{caminho_Empresa}\ROTINAS\Arquivos site\{data} - Empresa I.txt',
            'w') as site:
        site.write(f'<td class="td_year_title">2023</td>\n')
        site.write(f'				<td class="td_month">{teste(round(empItext.cell(row=92, column=3).value * 100, 2))}</td>\n')  #
        site.write(f'				<td class="td_month">{teste(round(empItext.cell(row=92, column=4).value * 100, 2))}</td>\n')  #
        site.write(f'				<td class="td_month">{teste(round(empItext.cell(row=92, column=5).value * 100, 2))}</td>\n')  #
        site.write(f'				<td class="td_month">{teste(round(empItext.cell(row=92, column=6).value * 100, 2))}</td>\n')  #
        site.write(f'				<td class="td_month">{teste(round(empItext.cell(row=92, column=7).value * 100, 2))}</td>\n')  #
        site.write(f'				<td class="td_month">{teste(round(empItext.cell(row=92, column=8).value * 100, 2))}</td>\n')
        site.write(f'				<td class="td_month">{teste(round(empItext.cell(row=92, column=9).value * 100, 2))}</td>\n')
        site.write(f'				<td class="td_month">{teste(round(empItext.cell(row=92, column=10).value * 100, 2))}</td>\n')
        site.write(f'				<td class="td_month">{teste(round(empItext.cell(row=92, column=11).value * 100, 2))}</td>\n')
        site.write(f'				<td class="td_month">{teste(round(empItext.cell(row=92, column=12).value * 100, 2))}</td>\n')
        site.write(f'				<td class="td_month">{teste(round(empItext.cell(row=92, column=13).value * 100, 2))}</td>\n')
        site.write(f'               <td class="td_month">{teste(round(empItext.cell(row=92, column=14).value * 100, 2))}</td>\n')
        site.write(
            f'				<td class="td_year">{round(empItext.cell(row=92, column=15).value * 100, 2)}%</td>\n')  # esse ano
        site.write(
            f'				<td class="td_inception">{round((empItext.cell(row=92, column=16).value) * 100, 2)}%</td>\n')  # acumulado
        site.write(
            f'				<td class="td_bench_year">{round((empItext.cell(row=94, column=15).value) * 100, 2)}%</td>\n')  # bench ano
        site.write(
            f'				<td class="td_bench_inception">{round((empItext.cell(row=94, column=16).value) * 100, 2)}%</td>\n')  # bench acumulado
        site.write('			</tr>\n')
        site.write('		  </tbody>\n')
        site.write('		  <tfoot>\n')
        site.write('			  <tr class="net">\n')
        site.write(
            f'				  <td colspan="6" class="net_worth">Património Líquido: {"R$  {:,.2f}".format(float(cart_f1_ont.pl))}</td>\n')
        site.write(
            f'				  <td colspan="6" class="net_worth_twelve">Património Líquido Médio(12m):  {"R$  {:,.2f}".format(sum(plsi) / float(len(plsi)))}</td>\n')
        site.write(
            f'				  <td colspan="5" class="quota">Valor da Cota: R$  {cart_f1_ont.cota}</td>\n')
        site.write('			  </tr>\n')
        site.write('			  <tr class="updated-date">\n')
        site.write(f'				  <td colspan="18">Atualizado em: {dia}/{mes}/{ano} 09:03</td>')

    with open(
            rf'C:\Users\{caminho_Empresa}\ROTINAS\Arquivos site\{data} - Empresa II.txt',
            'w') as site:
        site.write(f'<td class="td_year_title">2023</td>\n')
        site.write(f'				<td class="td_month">{teste(round(empIitext.cell(row=79, column=3).value * 100, 2))}</td>\n')  #
        site.write(f'				<td class="td_month">{teste(round(empIitext.cell(row=79, column=4).value * 100, 2))}</td>\n')  #
        site.write(f'				<td class="td_month">{teste(round(empIitext.cell(row=79, column=5).value * 100, 2))}</td>\n')  #
        site.write(f'				<td class="td_month">{teste(round(empIitext.cell(row=79, column=6).value * 100, 2))}</td>\n')  #
        site.write(f'				<td class="td_month">{teste(round(empIitext.cell(row=79, column=7).value * 100, 2))}</td>\n')  #
        site.write(f'				<td class="td_month">{teste(round(empIitext.cell(row=79, column=8).value * 100, 2))}</td>\n')
        site.write(f'				<td class="td_month">{teste(round(empIitext.cell(row=79, column=9).value * 100, 2))}</td>\n')
        site.write(f'				<td class="td_month">{teste(round(empIitext.cell(row=79, column=10).value * 100, 2))}</td>\n')
        site.write(f'				<td class="td_month">{teste(round(empIitext.cell(row=79, column=11).value * 100, 2))}</td>\n')
        site.write(f'				<td class="td_month">{teste(round(empIitext.cell(row=79, column=12).value * 100, 2))}</td>\n')
        site.write(f'				<td class="td_month">{teste(round(empIitext.cell(row=79, column=13).value * 100, 2))}</td>\n')
        site.write(f'				<td class="td_month">{teste(round(empIitext.cell(row=79, column=14).value * 100, 2))}</td>\n')  # esse mês
        site.write(
            f'				<td class="td_year">{round(empIitext.cell(row=79, column=15).value * 100, 2)}%</td>\n')  # esse ano
        site.write(
            f'				<td class="td_inception">{round((empIitext.cell(row=79, column=16).value) * 100, 2)}%</td>\n')  # acumulado
        site.write(
            f'				<td class="td_bench_year">{round((empIitext.cell(row=81, column=15).value) * 100, 2)}%</td>\n')  # bench ano
        site.write(
            f'				<td class="td_bench_inception">{round((empIitext.cell(row=81, column=16).value) * 100, 2)}%</td>\n')  # bench acumulado
        site.write('			</tr>\n')
        site.write('		  </tbody>\n')
        site.write('		  <tfoot>\n')
        site.write('			  <tr class="net">\n')
        site.write(
            f'				  <td colspan="6" class="net_worth">Património Líquido: {"R$  {:,.2f}".format(float(cart_f2_ont.pl))}</td>\n')
        site.write(
            f'				  <td colspan="6" class="net_worth_twelve">Património Líquido Médio(12m):  {"R$  {:,.2f}".format(sum(plsii) / float(len(plsii)))}</td>\n')
        site.write(
            f'				  <td colspan="5" class="quota">Valor da Cota: R$  {cart_f2_ont.cota}</td>\n')
        site.write('			  </tr>\n')
        site.write('			  <tr class="updated-date">\n')
        site.write(f'				  <td colspan="18">Atualizado em: {dia}/{mes}/{ano} 09:03</td>')
    outlook = win32.Dispatch('outlook.application')
    mail = outlook.CreateItem(0)
    mail.To = 'mail1'
    mail.CC = 'mail2'
    mail.Subject = f'Atualização do site - {dia}/{mes}/{ano}'
    mail.Body = f'Bom dia!\n\nSeguem os arquivos para atualização do site referentes ao dia {ontem_bonito}.\n\nAtt,'
    mail.Attachments.Add(
        rf'C:\Users\{caminho_Empresa}\ROTINAS\Arquivos site\{data} - Empresa I.txt')
    mail.Attachments.Add(
        rf'C:\Users\{caminho_Empresa}\ROTINAS\Arquivos site\{data} - Empresa II.txt')
    mail.Display()


def comando19():
    excel = win32.Dispatch('Excel.Application')
    wb = excel.Workbooks.Open(
        rf'C:\Users\{caminho_Empresa}\ROTINAS\Relatorios Diarios\Excel\{ontem}.xlsx')
    sheet = wb.Sheets[0]
    excel.visible = 1
    copyrange = sheet.Range('A1:Q99')
    copyrange.CopyPicture(Appearance=1, Format=2)
    ImageGrab.grabclipboard().save(
        rf'C:\Users\{caminho_Empresa}\ROTINAS\Relatorios Diarios\Empresa I\Relatório diário Empresa I - {ontem_arq}.pdf')
    sheet2 = wb.Sheets[1]
    copyrange2 = sheet2.Range('A1:Q99')
    copyrange2.CopyPicture(Appearance=1, Format=2)
    ImageGrab.grabclipboard().save(
        rf'C:\Users\{caminho_Empresa}\ROTINAS\Relatorios Diarios\Empresa II\Relatório diário Empresa II - {ontem_arq}.pdf')
    excel.Quit()
    outlook = win32.Dispatch('outlook.application')
    mail = outlook.CreateItem(0)
    mail.To = 'mail1'
    mail.CC = 'mail2'
    mail.Subject = f'Relatórios Diários - {dia}/{mes}/{ano}'
    mail.Body = f'Bom dia!\n\nSeguem os relatórios diários referentes ao dia {ontem_bonito}.\n\nAtt,'
    mail.Attachments.Add(
        rf'C:\Users\{caminho_Empresa}\ROTINAS\Relatorios Diarios\Empresa I\Relatório diário Empresa I - {ontem_arq}.pdf')
    mail.Attachments.Add(
        rf'C:\Users\{caminho_Empresa}\ROTINAS\Relatorios Diarios\Empresa II\Relatório diário Empresa II - {ontem_arq}.pdf')
    mail.Display()
    mail = outlook.CreateItem(0)
    mail.To = 'mail1'
    mail.CC = 'mail2'
    mail.Subject = f'POSIÇÃO - INVESTIDOR 1 - Empresa I FIC FIA'
    mail.Body = f'Bom dia!\n\nSegue posição para Investidor 1 referente ao dia {ontem_bonito}.\n\n\nAtt,'
    mail.Display()


def comando20():
    ops = []
    try:
        tabelamovs = xlrd.open_workbook(
            fr'C:\Users\{caminho_Empresa}\ROTINAS\Resgates - Caixa\Históricos\{ontem}.xls')
    except:
        caminhos = [os.path.join(
            rf'C:\Users\{caminho_Empresa}\ROTINAS\PASTA INPUT',
            nome) for nome in os.listdir(
            rf'C:\Users\{caminho_Empresa}\ROTINAS\PASTA INPUT')]
        movimentos = [arq for arq in caminhos if os.path.isfile(arq)]
        for i in movimentos:
            if 'Movimentos' in i:
                os.rename(i,
                          fr'C:\Users\{caminho_Empresa}\ROTINAS\Resgates - Caixa\Históricos\{ontem}.xls')
        tabelamovs = xlrd.open_workbook(
            fr'C:\Users\{caminho_Empresa}\ROTINAS\Resgates - Caixa\Históricos\{ontem}.xls')
    tabmovs = tabelamovs.sheet_by_index(0)
    for j in range(1, tabmovs.nrows):
        x = tabmovs.row_values(j)
        if len(str(x[18])) != 1:
            ops.append([x[4], x[11], x[3], x[1], x[18], x[5], x[7]])  # [operação, valor bruto, fundo resgatado, cotista, qtdcota, datamov, dataliq]
        else:
            ops.append([x[4], x[11], x[3], x[1], x[13], x[5], x[7]]) # [operação, valor bruto, fundo resgatado, cotista, qtdcota, datamov, dataliq]
    wb = openpyxl.load_workbook(
        rf'C:\Users\{caminho_Empresa}\ROTINAS\Resgates - Caixa\Template - Passivo Britech.xlsx')
    ws = wb.sheetnames
    planilha = wb[ws[0]]
    s = 2
    for f in ops:  # [operação, valor bruto, fundo resgatado, cotista, qtdcota, datamov, dataliq]
        if f[0] == 'A':
            tp = ''
        else:
            tp = 4
        if f[3] == 'EMP0012':
            fundo = 12
        else:
            fundo = 30
        planilha.cell(row=s, column=1).value = fundo # id do cotista
        planilha.cell(row=s, column=2).value = int(mudancas[f[2]])
        planilha.cell(row=s, column=3).value = transf_data(f[5])
        planilha.cell(row=s, column=4).value = transf_data(f[5])
        planilha.cell(row=s, column=5).value = transf_data(f[6])
        planilha.cell(row=s, column=6).value = mudancas[f[0]]
        planilha.cell(row=s, column=7).value = tp
        planilha.cell(row=s, column=8).value = float(f[4])
        planilha.cell(row=s, column=9).value = float(f[1]) / float(f[4])
        planilha.cell(row=s, column=10).value = f[1]
        planilha.cell(row=s, column=11).value = f[1]
        planilha.cell(row=s, column=12).value = 0
        planilha.cell(row=s, column=13).value = 0
        planilha.cell(row=s, column=14).value = 0
        planilha.cell(row=s, column=15).value = 0
        planilha.cell(row=s, column=16).value = 'Brasil'
        s += 1
    wb.save(
        rf'C:\Users\{caminho_Empresa}\ROTINAS\Resgates - Caixa\Upload - Britech\Passivo - {ontem}.xlsx')
    wb = openpyxl.load_workbook(
        rf'C:\Users\{caminho_Empresa}\ROTINAS\Resgates - Caixa\Template - Ativo Britech.xlsx')
    ws = wb.sheetnames
    planilha = wb[ws[0]]
    k = 2
    for f in ops:  # [operação, valor bruto, fundo resgatado, cotista, qtdcota, datamov, dataliq]
        if f[0] == 'A':
            tp = ''
        else:
            tp = 4
        planilha.cell(row=k, column=1).value = mudancas[f[3]]  # id do cotista
        planilha.cell(row=k, column=2).value = mudancas[f[2]]
        planilha.cell(row=k, column=3).value = transf_data(f[5])
        planilha.cell(row=k, column=4).value = transf_data(f[5])
        planilha.cell(row=k, column=5).value = transf_data(f[6])
        planilha.cell(row=k, column=6).value = mudancas[f[0]]
        planilha.cell(row=k, column=7).value = tp
        planilha.cell(row=k, column=8).value = float(f[4])
        planilha.cell(row=k, column=9).value = float(f[1])/float(f[4])
        planilha.cell(row=k, column=10).value = f[1]
        planilha.cell(row=k, column=11).value = f[1]
        planilha.cell(row=k, column=12).value = 0
        planilha.cell(row=k, column=13).value = 0
        planilha.cell(row=k, column=14).value = 0
        planilha.cell(row=k, column=15).value = 0
        planilha.cell(row=k, column=16).value = 'Brasil'
        k += 1
    wb.save(
        rf'C:\Users\{caminho_Empresa}\ROTINAS\Resgates - Caixa\Upload - Britech\Ativo- {ontem}.xlsx')
    tk.messagebox.showinfo('Gerar Caixa',
                           'Os arquivos foram feitos e se encontram na pasta "Resgates - Caixa"\nFaça o upload deles nas áreas respectivas.')


def comando21():
    caminhos = [os.path.join(
        rf'C:\Users\{caminho_Empresa}\ROTINAS\Zeragem - Britech\Operacoes admin\{ontem}',
        nome) for nome in os.listdir(
        rf'C:\Users\{caminho_Empresa}\ROTINAS\Zeragem - Britech\Operacoes admin\{ontem}')]
    movimentos = [arq for arq in caminhos if os.path.isfile(arq) and '.xls' in arq]
    totalfd = pd.DataFrame(columns=['Data', 'Op.', 'Fundo', 'Valor da Cota', 'Valor Bruto', 'Fundo'])
    totalrf = pd.DataFrame(columns=['Data', 'Op.', 'Fundo', 'Vlr. Líquido', 'PU', 'Quantidade',
                                    'Cód. CETIP/SELIC'])
    for i in movimentos:
        if 'Fundo' in i:
            if 'cod2' in i:
                fd = 'cod2'
            elif 'cod1' in i:
                fd = 'cod1'
            elif 'cod4' in i:
                fd = 'cod4'
            else:
                fd = 'cod3'
            k = pd.read_excel(i, 0, skiprows=5)
            k.insert(0, 'Fundo', fd)
            a = k[k['Fundo'] == 'BEM FI RF SIMPL TPF']
            totalfd = pd.concat([totalfd, a], join='inner')
        elif 'RendaFixa' in i:
            k = pd.read_excel(i, 0, skiprows=6).fillna(method='bfill')
            k = k.drop(
                [2, len(list(k['Unnamed: 0'])) - 1, len(list(k['Unnamed: 0'])) - 2, len(list(k['Unnamed: 0'])) - 3,
                 len(list(k['Unnamed: 0'])) - 4,
                 len(list(k['Unnamed: 0'])) - 5, len(list(k['Unnamed: 0'])) - 6])
            p1 = k.iloc[::2]
            p1.columns = p1.iloc[0]
            p1 = p1.drop(0).reset_index().drop(columns=[a for a in list(p1) if
                                                        a not in ['Data', 'Op.', 'Fundo', 'Vlr. Líquido', 'PU',
                                                                  'Quantidade', 'Cód. CETIP/SELIC']])
            p2 = k.iloc[1::2]
            p2.columns = p2.iloc[0]
            p2 = p2.drop(1).reset_index().drop(columns=[a for a in list(p2) if
                                                        a not in ['Data', 'Op.', 'Fundo', 'Vlr. Líquido', 'PU',
                                                                  'Quantidade', 'Cód. CETIP/SELIC']])
            p2 = p2.drop(columns=['index'])
            p1 = p1.drop(columns=['index'])
            comp = p1.join(p2, how='left')
            comp['PU'] = comp['Vlr. Líquido'] / comp['Quantidade']
            if 'cod3' in i:
                comp.insert(0, 'Fundo', 'cod3')
            elif 'cod4' in i:
                comp.insert(0, 'Fundo', 'cod4')
            elif 'cod1' in i:
                comp.insert(0, 'Fundo', 'cod1')
            else:
                comp.insert(0, 'Fundo', 'cod2')
            totalrf = pd.concat([totalrf, comp], join='inner')
    if len(totalrf['Data']) > 0:
        totalrf = totalrf[totalrf['Op.'] == 'C']
        with open(
                fr'C:\Users\{caminho_Empresa}\ROTINAS\Relatorios Diarios\Bases de dados\Arquivos CDI\{ontem}.txt',
                'r', encoding='latin-1') as cd:
            a = cd.readlines()[2].split(";", 1000)
            taxa = float(a[1].replace(',', '.'))
            fat_cdi = float(a[2].replace(',', '.'))
        titulos = {950199: 2245825, 210100: 2245806, 100000: 2245838, 760199: 2245814}
        wb = openpyxl.load_workbook(
            rf'C:\Users\{caminho_Empresa}\ROTINAS\Zeragem - Britech\Template RF - Zeragem.xlsx')
        ws = wb.sheetnames
        planilha = wb[ws[0]]
        s = 2
        for k in zip(list(totalrf['Fundo']), list(totalrf['Data']), list(totalrf['Op.']),
                     list(totalrf['Vlr. Líquido']), list(totalrf['PU']), list(totalrf['Quantidade']),
                     list(totalrf['Cód. CETIP/SELIC'])):
            if mudancas[k[0]] == 387894:
                conta = 1
            elif mudancas[k[0]] == 454109:
                conta = 156
            elif mudancas[k[0]] == 652091:
                conta = 164
            else:
                conta = 157
            if pd.to_datetime(k[1], format='%d/%m/%Y') == 5:
                d1 = pd.to_datetime(k[1], format='%d/%m/%Y') + datetime.timedelta(days=3)
            else:
                d1 = pd.to_datetime(k[1], format='%d/%m/%Y') + datetime.timedelta(days=1)
            planilha.cell(row=s, column=1).value = mudancas[k[0]]
            planilha.cell(row=s, column=2).value = titulos[k[6]]
            planilha.cell(row=s, column=3).value = k[1]
            planilha.cell(row=s, column=4).value = 'CompraRevenda'
            planilha.cell(row=s, column=5).value = k[5]
            planilha.cell(row=s, column=6).value = k[4]
            planilha.cell(row=s, column=7).value = k[3]
            planilha.cell(row=s, column=8).value = taxa
            planilha.cell(row=s, column=10).value = muda_data(str(d1).replace('-', ''))
            planilha.cell(row=s, column=11).value = taxa
            planilha.cell(row=s, column=12).value = k[4] * fat_cdi
            planilha.cell(row=s, column=13).value = (k[4] * fat_cdi) * k[5]
            planilha.cell(row=s, column=32).value = conta
            s += 1
        wb.save(
            rf'C:\Users\{caminho_Empresa}\ROTINAS\Zeragem - Britech\Importação diária\{ontem} - RF.xls')
    if len(totalfd['Fundo']) > 0:
        wb = openpyxl.load_workbook(
            rf'C:\Users\{caminho_Empresa}\ROTINAS\Zeragem - Britech\Template Fundo - Zeragem.xlsx')
        ws = wb.sheetnames
        planilha = wb[ws[0]]
        totalfd['Operação'] = [2 if a == 'R' else 4 for a in list(totalfd['Op.'])]
        totalfd['TipoResg'] = [3 if a == 2 else '' for a in list(totalfd['Operação'])]
        s = 2
        for k in zip(list(totalfd['Fundo']), [f'{a[0:6]}20{a[6:8]}' for a in list(totalfd['Data'])],
                     list(totalfd['Operação']),
                     list(totalfd['Valor Bruto']), list(totalfd['TipoResg']), list(totalfd['Valor da Cota'])):
            if mudancas[k[0]] == 387894:
                conta = 1
            elif mudancas[k[0]] == 454109:
                conta = 156
            elif mudancas[k[0]] == 652091:
                conta = 164
            else:
                conta = 157
            planilha.cell(row=s, column=1).value = mudancas[k[0]]
            planilha.cell(row=s, column=2).value = 207632
            planilha.cell(row=s, column=3).value = k[1]
            planilha.cell(row=s, column=4).value = k[1]
            planilha.cell(row=s, column=5).value = k[1]
            planilha.cell(row=s, column=6).value = k[2]
            planilha.cell(row=s, column=7).value = k[4]
            planilha.cell(row=s, column=8).value = k[3] / k[5]
            planilha.cell(row=s, column=9).value = k[5]
            planilha.cell(row=s, column=10).value = k[3]
            planilha.cell(row=s, column=11).value = k[3]
            planilha.cell(row=s, column=12).value = 0
            planilha.cell(row=s, column=13).value = 0
            planilha.cell(row=s, column=14).value = 0
            planilha.cell(row=s, column=15).value = 0
            planilha.cell(row=s, column=16).value = 'selic'
            planilha.cell(row=s, column=26).value = conta
            s += 1
        wb.save(
            rf'C:\Users\{caminho_Empresa}\ROTINAS\Zeragem - Britech\Importação diária\{ontem} - Fundo.xlsx')
        wbz = openpyxl.load_workbook(
            rf'C:\Users\{caminho_Empresa}\ROTINAS\Zeragem - Britech\Template Cota Zeragem.xlsx')
        wsz = wbz.sheetnames
        planilhaz = wbz[wsz[0]]
        planilhaz.cell(row=2, column=1).value = 207632
        planilhaz.cell(row=2, column=2).value = f'{ontem[6:8]}/{ontem[4:6]}/{ontem[0:4]}'
        planilhaz.cell(row=2, column=3).value = list(totalfd['Valor da Cota'])[0]
        planilhaz.cell(row=2, column=4).value = 0
        planilhaz.cell(row=2, column=5).value = 0
        planilhaz.cell(row=2, column=6).value = 0
        planilhaz.cell(row=2, column=7).value = 0
        wbz.save(
            rf'C:\Users\{caminho_Empresa}\ROTINAS\Zeragem - Britech\Cotações - BEM FI RF\Zeragem - {ontem}.xlsx')
    tk.messagebox.showinfo('Zeragem Britech',
                           'Os arquivos estão no caminho respectivo. Faça o upload deles na Britech.')


def comando22():
    ativos_offshore = ['MELI US', 'XP US', 'OXY US', 'CVX US', 'XLE US', 'XOP US', 'USO US']
    chromeoptions = webdriver.ChromeOptions()
    prefs = {'safebrowsing.enabled': 'false'}
    chromeoptions.add_experimental_option("prefs", prefs)
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
    time.sleep(2)
    driver.implicitly_wait(10)
    try:
        xlrd.open_workbook(
            fr'C:\Users\{caminho_Empresa}\ROTINAS\Balanco\IMA-B\{ontem}.xls')
        aj = 1
    except:
        driver.maximize_window()
        driver.get("https://www.anbima.com.br/informacoes/ima/arqs/ima_completo.xls")
        time.sleep(5)
        os.rename(fr'C:\Users\{os.getlogin()}\Downloads\ima_completo.xls',
                  fr'C:\Users\{caminho_Empresa}\ROTINAS\Balanco\IMA-B\{ontem}.xls')
        aj = 0
    driver.get(
        "https://www.b3.com.br/pt_br/market-data-e-indices/servicos-de-dados/market-data/consultas/mercado-de-derivativos/indicadores/indicadores-financeiros/")
    driver.switch_to.frame("bvmf_iframe")
    dolarrr = driver.find_element(By.XPATH,
                                  '//*[@id="divContainerIframeB3"]/form/div/div/div/div/div[2]/div[3]/div/div/h4').text
    dolarr = dolarrr.replace(' (R$/US$)', '')
    dolar = float(dolarr.replace(',', '.'))
    driver.quit()
    ima = xlrd.open_workbook(
        fr'C:\Users\{caminho_Empresa}\ROTINAS\Balanco\IMA-B\{ontem}.xls')
    tab1imab = ima.sheet_by_index(6)
    imab = tab1imab.cell_value(4, 19) / 100
    fat_imab = (1 + imab) ** (1 / 252)
    with open(
            fr'C:\Users\{caminho_Empresa}\ROTINAS\Relatorios Diarios\Bases de dados\IPCA_atual.txt',
            'r') as ip:
        ipc = ip.readline()
        ipcc = ipc.replace('IPCA Projetado Atual: ', '')
        ipca = float(ipcc.replace(',', '.'))
    fat_ipca = (1 + ipca) ** (1 / 22)
    bench = fat_ipca * fat_imab
    k = pd.read_excel(
        rf'C:\Users\{caminho_Empresa}\ROTINAS\Relatorios Diarios\Bases de dados\Hist Benchmark.xlsx',
        0)
    listabench = list(k['Cotacao'])
    bench_novo = ((1 + listabench[len(listabench) - aj - 1]) * bench) - 1
    wbd = openpyxl.load_workbook(
        rf'C:\Users\{caminho_Empresa}\ROTINAS\Cotações Britech\Template índices.xlsx')
    wsd = wbd.sheetnames
    planilhad = wbd[wsd[0]]
    planilhad.cell(row=2, column=1).value = f'{ontem[6:8]}/{ontem[4:6]}/{ontem[0:4]}'
    planilhad.cell(row=2, column=2).value = 99
    planilhad.cell(row=2, column=3).value = dolar
    #planilhad.cell(row=2, column=4).value = f'{ontem[6:8]}/{ontem[4:6]}/{ontem[0:4]}'
    planilhad.cell(row=2, column=5).value = 2
    #planilhad.cell(row=2, column=6).value = f'{ontem[6:8]}/{ontem[4:6]}/{ontem[0:4]}'
    planilhad.cell(row=3, column=1).value = f'{ontem[6:8]}/{ontem[4:6]}/{ontem[0:4]}'
    planilhad.cell(row=3, column=2).value = 2017
    planilhad.cell(row=3, column=3).value = bench_novo
    #planilhad.cell(row=3, column=4).value = f'{ontem[6:8]}/{ontem[4:6]}/{ontem[0:4]}'
    planilhad.cell(row=3, column=5).value = 2
    #planilhad.cell(row=3, column=6).value = f'{ontem[6:8]}/{ontem[4:6]}/{ontem[0:4]}' #Linhas tiradas porque não precisa de data de vigência se a série é diária.
    wbd.save(
        rf'C:\Users\{caminho_Empresa}\ROTINAS\Cotações Britech\Índices\Índices - {ontem}.xlsx')
    cotacoes = pd.DataFrame()
    for i in ativos_offshore:
        ativo = yf.Ticker(f'{i.replace(" US", "")}')
        hist = ativo.history(period='1d')
        cotacoes = pd.concat([cotacoes, hist])
    cotacoes = cotacoes.reset_index()
    wb = openpyxl.load_workbook(
        rf'C:\Users\{caminho_Empresa}\ROTINAS\Cotações Britech\Template ativos Offshore.xlsx')
    ws = wb.sheetnames
    planilha = wb[ws[0]]
    lista_planilha = [[f'{ontem[6:8]}/{ontem[4:6]}/{ontem[0:4]}' for a in ativos_offshore], ativos_offshore,
                      list(cotacoes['Close']), list(cotacoes['Close'])]
    s = 1
    for f in lista_planilha:
        x = 0
        for k in f:
            planilha.cell(row=2 + x, column=s).value = k
            x += 1
        s += 1
    try:  # esse bloco tenta pegar a precificação de opções caso a Britech não pegue.
        try:
            opcoes = cart_m1_ont.opcoes
        except:
            opcoes = cart_m2_ont.opcoes
        if opcoes != [['-', '-', '-', '-', '-', '-', '-', '-']]:
            listaopccoes = [[f'{ontem[6:8]}/{ontem[4:6]}/{ontem[0:4]}' for a in opcoes], [a[0] for a in opcoes], [a[3] for a in opcoes], [a[3] for a in opcoes]]
            ss = 1
            for a in listaopccoes:
                xx = 0
                for k in a:
                    planilha.cell(row=(len(lista_planilha[0])+2) + xx, column=ss).value = k
                    xx += 1
                ss += 1
    except:
        pass
    wb.save(
        rf'C:\Users\{caminho_Empresa}\ROTINAS\Cotações Britech\Ativos Offshore\Ativos offshore - {ontem}.xlsx')
    tk.messagebox.showinfo('Cotações Britech',
                           'Os arquivos estão no caminho respectivo. Faça o upload deles na Britech.')


def comando23():
    def aux(a):
        k = str(pd.to_datetime(a, format='%d/%m/%Y')).replace(' 00:00:00', '')
        return k.replace('-', '')

    dados_23 = []
    prompt23 = tk.Tk()
    c23 = tk.Canvas(prompt23, width=400, height=280, relief='raised', bg=preto_Empresa)
    c23.pack()
    v231 = tk.StringVar(prompt23)
    v232 = tk.StringVar(prompt23)
    l23 = tk.Label(prompt23, text='Relatório de Performance Attribuition', font=('helvetica', 14), bg=preto_Empresa, fg=amarelo_Empresa)
    c23.create_window(200, 25, window=l23)
    l231 = tk.Label(prompt23, text='Digite a data de início:  (YYYYMMDD)',
                   font=('helvetica', 10), bg=preto_Empresa, fg=amarelo_Empresa)
    c23.create_window(200, 85, window=l231)
    ent232 = tk.Entry(prompt23, textvariable=v231)
    c23.create_window(200, 110, window=ent232)

    l232 = tk.Label(prompt23, text='Digite a data de fim:  (YYYYMMDD)',
                   font=('helvetica', 10), bg=preto_Empresa, fg=amarelo_Empresa)
    c23.create_window(200, 155, window=l232)
    ent232 = tk.Entry(prompt23, textvariable=v232)
    c23.create_window(200, 180, window=ent232)

    def pegainput23():
        global w231, w232
        w231 = v231.get()
        w232 = v232.get()
        dados_23.append([w231, w232])
        bot231.config(text='Pronto!', bg=amarelo_Empresa, fg=preto_Empresa)

    bot231 = tk.Button(prompt23, text='Registrar >', command=pegainput23, bg=branco_Empresa, fg=preto_Empresa,
                     font=('helvetica', 9, 'bold'))
    c23.create_window(200, 250, window=bot231)
    bot232 = tk.Button(prompt23, text='Concluir!', command=lambda: [f() for f in [prompt23.destroy, root1.quit]],
                      bg=amarelo_Empresa, fg=preto_Empresa,
                      font=('helvetica', 9, 'bold'))
    c23.create_window(300, 250, window=bot232)
    prompt23.mainloop()

    tk.messagebox.showinfo('Relatório de Performance Attribuition',
                           'Esse processo pode demorar alguns instantes. Por favor, aguarde.')
    dados_funcao = dados_23[0]
    data_inicio = dados_funcao[0]
    data_fim = dados_funcao[1]
    if data_inicio[0:4] == '2022':
        inicio_ano = '20211231'
    elif data_inicio[0:4] == '2023':
        inicio_ano = '20221230'
    elif data_inicio[0:4] == '2024':
        inicio_ano = '20231229'
    else:
        inicio_ano = '20241231'  #aqui, para relatórios puxados depois do ano de 2025, deve-se acrescer linhas.
    periodo = pd.DataFrame()
    acumulado = pd.DataFrame(columns=['Data', 'CTR', 'Ticker', 'Empresa'])
    for jj in [inicio_ano, data_inicio]:
        brutao = pd.read_excel(
            fr'C:\Users\{caminho_Empresa}\ROTINAS\Perfromance Atribuittion\Empresa II\Rentabilidade por ativo - Empresa II.xlsx',
            0)
        brutao['Indice'] = pd.to_datetime(brutao['Data'], format='%d/%m/%Y')
        brutin = brutao[brutao['Indice'] > pd.to_datetime(jj)]
        rent_brut = brutin[brutin['Indice'] <= pd.to_datetime(data_fim)]
        rent_brut.loc[(rent_brut['Tipo Ativo'] == 'Op.Renda Fixa - Tributado') & (
                    rent_brut['Rendimento Líquido'] == 0),
                      'Rendimento Líquido'] = rent_brut['Patrimonio Final Líquido - Carteira'] * -1
        rent_brut = rent_brut.drop(columns=[a for a in list(rent_brut) if
                                            a not in ['Data', 'Indice', 'Id Ativo', 'Rendimento Líquido',
                                                      'Vl.Rendas - Carteira']])
        rent_brut = rent_brut.groupby(['Data', 'Id Ativo'], as_index=False).sum()
        rent = pd.pivot_table(rent_brut, columns='Id Ativo', values='Rendimento Líquido', index='Data').fillna(
            0).reset_index()
        lista = [float(Carteira(jj, 'f2').pl)]
        for i in [float(Carteira(aux(a), 'f2').pl) for a in list(rent['Data'])]:
            lista.append(i)
        lista.pop(-1)
        rent['Patrimônio Líquido do Fundo'] = lista
        rent['Variação da Cota'] = [
            float(Carteira(data_fim, 'f2').cota) / float(Carteira(aux(a), 'f2').cota) for a in
            list(rent['Data'])]
        for k in [a for a in list(rent) if a not in ['Data', 'Patrimônio Líquido do Fundo', 'Variação da Cota']]:
            rent[k] = rent[k] / rent['Patrimônio Líquido do Fundo']
        dict_rents = {}
        for k in [a for a in list(rent) if a not in ['Data', 'Patrimônio Líquido do Fundo', 'Variação da Cota']]:
            rent[k] = rent[k] * rent['Variação da Cota']
        for k in [a for a in list(rent) if a not in ['Data', 'Patrimônio Líquido do Fundo', 'Variação da Cota']]:
            dict_rents[k] = sum(list(rent[k]))
        rents = pd.DataFrame({'Ticker': list(dict_rents), 'CTR': [dict_rents[a] for a in dict_rents]})
        rents['Empresa'] = rents['Ticker'].str[:4]
        rents.insert(0, 'CTR Opções', 0)
        rents.loc[(rents['Ticker'].str.len() > 6) & (rents['Ticker'] != 'MELI US'), 'CTR Opções'] = rents['CTR']
        rents['CTR Ações'] = rents['CTR'] - rents['CTR Opções']
        if jj == data_inicio:
            periodo = pd.concat([periodo, rents])
        else:
            acumlado = pd.concat([acumulado, rents], join='inner').rename(columns={'CTR': 'CTR acumulado'})
            acumulado = acumlado.groupby('Empresa', as_index=False).sum()
    carteirao = pd.DataFrame(columns=['Data', 'Ticker', 'Preco', '% do PL'])
    for i in range(int(data_inicio), int(data_fim)+1):
        try:
            carteirao = pd.concat([carteirao, Carteira(i, 'm2').pega_carteira(5)], join='inner')
        except:
            pass
    carteirao['Empresa'] = carteirao['Ticker'].str[0:4]
    dict_media = {}
    dict_var = {}
    for i in set(list(carteirao['Empresa'])):
        acao = carteirao[carteirao['Empresa'] == i].reset_index()
        dict_var[i] = acao.loc[len(acao['Preco']) - 1, 'Preco'] / acao.loc[0, 'Preco'] - 1
        acao['% do PL'] = acao['% do PL'].astype(float)
        acao = acao.groupby(['Data', 'Empresa'], as_index=False).sum()
        dict_media[i] = acao['% do PL'].mean()
    periodo = periodo.groupby('Empresa', as_index=False).sum()
    final = acumulado.merge(periodo, on='Empresa', how='left').fillna(0)
    final = final.drop(final.loc[final['Empresa'].isin([a for a in list(final['Empresa']) if '2' in a])].index)
    vars = pd.DataFrame({'Empresa':dict_var.keys(), 'Variação mensal':dict_var.values()}).merge(pd.DataFrame({'Empresa':dict_media.keys(), 'Concentração média':dict_media.values()}), how='left')
    final = final.merge(vars, how='left').fillna(0)
    # maiores compras e vendas por % do PL
    ops = pd.read_csv(
        fr'C:\Users\{caminho_Empresa}\ROTINAS\Perfromance Atribuittion\Empresa II\Operações - Empresa II.csv',
        encoding='latin-1', sep=';', skiprows=5).drop(0).iloc[:-2, :]
    ops = ops.drop(
        columns=[a for a in list(ops) if 'Unnamed' in a] + ['Valor IR', 'Corretagem', 'Taxas', 'Valor Líquido'])
    ops = ops[(pd.to_datetime(ops['Data'], format='%d/%m/%Y') >= pd.to_datetime(data_inicio)) & (
                pd.to_datetime(ops['Data'], format='%d/%m/%Y') <= pd.to_datetime(data_fim))]
    ops.loc[ops['Quantidade'] < 100, 'Quantidade'] = ops['Quantidade'] * 1000
    for a in ['Preço', 'Valor']:
        ops[a] = ops[a].str.replace('.', '', regex=True)
        ops[a] = ops[a].str.replace(',', '.', regex=True)
        ops[a] = ops[a].str.replace(')', '', regex=True)
        ops[a] = ops[a].str.replace('(', '-', regex=True)
        ops[a] = ops[a].astype(float)
    ops = ops.groupby(['Data', 'Código'], as_index=False).sum()
    ops['Preço médio'] = ops['Valor'] / ops['Quantidade'] * -1
    top_5 = ops.sort_values(by='Valor', ascending=True).head(5)
    top_5['PL'] = [float(Carteira(aux(a), 'f2').pl) for a in list(top_5['Data'])]
    top_5['% do PL'] = top_5['Valor'] / top_5['PL'] * -1
    # pegando dados informativos de rentabilidades:
    meses = {'01':'Jan', '02':'Fev', '03':'Mar','04':'Abr', '05':'Mai', '06':'Jun', '07':'Jul','08':'Ago', '09':'Set',
             '10':'Out', '11':'Nov', '12':'Dez'}
    exc = openpyxl.load_workbook(
            rf'C:\Users\{caminho_Empresa}\ROTINAS\Relatorios Diarios\Excel\{data_fim}.xlsx',
            data_only=True)
    exce = exc.sheetnames
    excel = exc[exce[1]]
    cota = excel.cell(row=79, column=2+int(data_fim[4:6])).value
    cota_acum = excel.cell(row=79, column=15).value
    ibov_acum = excel.cell(row=81, column=15).value
    cdi_acum = excel.cell(row=82, column=15).value
    exc.close()
    # passando pra Planilha:
    for item in ['CTR acumulado', 'Concentração média', 'CTR']:
        final = final.sort_values(item, ascending=False)
        wb = openpyxl.load_workbook(
            rf'C:\Users\{caminho_Empresa}\ROTINAS\Perfromance Atribuittion\Template.xlsx')
        ws = wb.sheetnames
        planilha = wb[ws[0]]
        planilha.cell(row=3, column=6).value = f'Relatório de Contribuição de Performance - {data_fim[4:6]}/{data_fim[2:4]}'
        planilha.cell(row=8, column=5).value = f'{data_fim[4:6]} - {data_fim[2:4]}'
        planilha.cell(row=8, column=6).value = cota
        planilha.cell(row=8, column=7).value = cota_acum
        planilha.cell(row=8, column=8).value = ibov_acum
        planilha.cell(row=8, column=9).value = cdi_acum
        lista_colunas = ['Variação mensal','Concentração média', 'Empresa', 'CTR Ações', 'CTR Opções', 'CTR', 'CTR acumulado']
        ss = 11
        for f in ['Código', '% do PL', 'Preço']:
            zz = 28
            for k in list(top_5[f]):
                planilha.cell(row=zz, column=ss).value = k
                zz += 1
            ss += 1
        s = 2
        for f in lista_colunas:
            z = 12
            for k in list(final[f]):
                planilha.cell(row=z, column=s).value = k
                z += 1
            s += 1
        wb.save(
        fr'C:\Users\{caminho_Empresa}\ROTINAS\Perfromance Atribuittion\Enviados\{data_inicio} - {data_fim}, {item}.xlsx')
    for item in ['CTR acumulado', 'Concentração média', 'CTR']:
        excel = win32.Dispatch('Excel.Application')
        wbb = excel.Workbooks.Open(fr'C:\Users\{caminho_Empresa}\ROTINAS\Perfromance Atribuittion\Enviados\{data_inicio} - {data_fim}, {item}.xlsx')
        sheet = wbb.Sheets[0]
        excel.visible = 1
        copyrange = sheet.Range(f'A1:N40')
        copyrange.CopyPicture(Appearance=1, Format=2)
        ImageGrab.grabclipboard().save(
                fr'C:\Users\{caminho_Empresa}\ROTINAS\Perfromance Atribuittion\Enviados\{data_inicio} - {data_fim}, por {item}.pdf')
        excel.Quit()
        time.sleep(2)
    outlook = win32.Dispatch('outlook.application')
    mail = outlook.CreateItem(0)
    mail.To = 'mail1'
    mail.CC = 'mail2'
    mail.Subject = f'Performance Attribuition - Empresa II'
    mail.Body = f'Boa tarde! \n\nSeguem os relatórios de Performance Contribuittion de {data_inicio[6:8]}/{data_inicio[4:6]}/{data_inicio[0:4]} a {data_fim[6:8]}/{data_fim[4:6]}/{data_fim[0:4]}. \n\nAtt,'
    attachment1 = fr'C:\Users\{caminho_Empresa}\ROTINAS\Perfromance Atribuittion\Enviados\{data_inicio} - {data_fim}, por CTR acumulado.pdf'
    attachment2 = fr'C:\Users\{caminho_Empresa}\ROTINAS\Perfromance Atribuittion\Enviados\{data_inicio} - {data_fim}, por CTR.pdf'
    attachment3 = fr'C:\Users\{caminho_Empresa}\ROTINAS\Perfromance Atribuittion\Enviados\{data_inicio} - {data_fim}, por Concentração média.pdf'
    mail.Attachments.Add(attachment1)
    mail.Attachments.Add(attachment2)
    mail.Attachments.Add(attachment3)
    mail.Display(True)
    tk.messagebox.showinfo('Relatório de Performance Attribuition',
                           'O relatório foi gerado com sucesso e já encontra-se na pasta.')


root1 = tk.Tk()
telacom = tk.Canvas(root1, width=1300, height=680, relief='raised', bg=preto_Empresa)
telacom.pack()
proc = tk.Label(root1, text='Processos - Empresa', font=('helvetica', 14), bg=preto_Empresa, fg=amarelo_Empresa)
telacom.create_window(675, 25, window=proc)

com1 = tk.Button(root1, text='Resgates Programados', padx=28, pady=20, command=comando1, bg=amarelo_Empresa, fg=preto_Empresa,
                 font=('helvetica', 9, 'bold'))
com2 = tk.Button(root1, text='Mov. Upload no Custódia', padx=40, pady=20, command=comando2, bg=preto_Empresa,
                 fg=amarelo_Empresa,
                 font=('helvetica', 9, 'bold'))
com3 = tk.Button(root1, text='Mov. Envio de email', padx=50, pady=20, command=comando3, bg=preto_Empresa, fg=amarelo_Empresa,
                 font=('helvetica', 9, 'bold'))
com4 = tk.Button(root1, text='Boletagem Renda Variável', padx=40, pady=20, command=comando4, bg=preto_Empresa,
                 fg=amarelo_Empresa,
                 font=('helvetica', 9, 'bold'))
com5 = tk.Button(root1, text='Boletagem câmbio', padx=50, pady=20, command=comando5, bg=preto_Empresa, fg=amarelo_Empresa,
                 font=('helvetica', 9, 'bold'))
com6 = tk.Button(root1, text='Boletagem Aluguel BTC', padx=50, pady=20, command=comando6, bg=preto_Empresa, fg=amarelo_Empresa,
                 font=('helvetica', 9, 'bold'))
com7 = tk.Button(root1, text='Gerar Caixa', padx=70, pady=20, command=comando7, bg=preto_Empresa, fg=amarelo_Empresa,
                 font=('helvetica', 9, 'bold'))
com8 = tk.Button(root1, text='Cadastrar Cotista', padx=50, pady=20, command=comando8, bg=preto_Empresa, fg=amarelo_Empresa,
                 font=('helvetica', 9, 'bold'))
com9 = tk.Button(root1, text='Relatório Cotações - Envio', padx=35, pady=20, command=comando9, bg=preto_Empresa,
                 fg=amarelo_Empresa,
                 font=('helvetica', 9, 'bold'))
com10 = tk.Button(root1, text='Relatório Cotações - Alimenta base', padx=10, pady=20, command=comando10, bg=preto_Empresa,
                  fg=amarelo_Empresa, font=('helvetica', 9, 'bold'))
com11 = tk.Button(root1, text='Alocar Arquivos', padx=50, pady=20, command=comando11, bg=branco_Empresa, fg=preto_Empresa,
                  font=('helvetica', 9, 'bold'))
com12 = tk.Button(root1, text='Relatório de BTC', padx=50, pady=20, command=comando12, bg=amarelo_Empresa, fg=preto_Empresa,
                  font=('helvetica', 9, 'bold'))
com13 = tk.Button(root1, text='Relatório diário e envio cotas', padx=15, pady=20, command=comando13, bg=amarelo_Empresa,
                  fg=preto_Empresa, font=('helvetica', 9, 'bold'))
com14 = tk.Button(root1, text='Taxa ADM e Rebates', padx=50, pady=20, command=comando14, bg=preto_Empresa,
                  fg=amarelo_Empresa,
                  font=('helvetica', 9, 'bold'))
com15 = tk.Button(root1, text='Rotinas Mensais', padx=50, pady=20, command=comando15, bg=preto_Empresa, fg=amarelo_Empresa,
                  font=('helvetica', 9, 'bold'))
com16 = tk.Button(root1, text='Relatório de Risco', padx=50, pady=20, command=comando16, bg=preto_Empresa, fg=amarelo_Empresa,
                  font=('helvetica', 9, 'bold'))
com17 = tk.Button(root1, text='Batimento da prévia', padx=35, pady=20, command=comando17, bg=amarelo_Empresa, fg=preto_Empresa,
                  font=('helvetica', 9, 'bold'))
com18 = tk.Button(root1, text='Envio arquivos site', padx=43, pady=20, command=comando18, bg=amarelo_Empresa, fg=preto_Empresa,
                  font=('helvetica', 9, 'bold'))
com19 = tk.Button(root1, text='Envio Email rel. diá. caso erro', padx=30, pady=20, command=comando19, bg=preto_Empresa,
                  fg=amarelo_Empresa,
                  font=('helvetica', 9, 'bold'))
com20 = tk.Button(root1, text='Registro op. de caixa Britech', padx=20, pady=20, command=comando20, bg=preto_Empresa,
                  fg=amarelo_Empresa, font=('helvetica', 9, 'bold'))
com21 = tk.Button(root1, text='Zeragem - Britech', padx=45, pady=20, command=comando21, bg=amarelo_Empresa, fg=preto_Empresa,
                  font=('helvetica', 9, 'bold'))
com22 = tk.Button(root1, text='Cotações - Britech', padx=50, pady=20, command=comando22, bg=amarelo_Empresa, fg=preto_Empresa,
                  font=('helvetica', 9, 'bold'))
com23 = tk.Button(root1, text='Rel. Performance Attribuition', padx=45, pady=20, command=comando23, bg=preto_Empresa,
                  fg=amarelo_Empresa, font=('helvetica', 9, 'bold'))
com24 = tk.Button(root1, text='Relatório variações', padx=50, pady=20, command=comando20, bg=preto_Empresa, fg=amarelo_Empresa,
                  font=('helvetica', 9, 'bold'))


telacom.create_window(200, 100, window=com11)
telacom.create_window(200, 200, window=com13)
telacom.create_window(200, 300, window=com17)
telacom.create_window(200, 400, window=com18)
telacom.create_window(200, 500, window=com12)
telacom.create_window(200, 600, window=com1)
telacom.create_window(500, 100, window=com22)
telacom.create_window(500, 200, window=com21)
telacom.create_window(500, 300, window=com2)
telacom.create_window(500, 400, window=com3)
telacom.create_window(500, 500, window=com4)
telacom.create_window(500, 600, window=com6)
telacom.create_window(800, 100, window=com14)
telacom.create_window(800, 200, window=com7)
telacom.create_window(800, 300, window=com9)
telacom.create_window(800, 400, window=com8)
telacom.create_window(800, 500, window=com15)
telacom.create_window(800, 600, window=com23)
telacom.create_window(1100, 100, window=com10)
telacom.create_window(1100, 200, window=com16)
telacom.create_window(1100, 300, window=com5)
telacom.create_window(1100, 400, window=com19)
telacom.create_window(1100, 500, window=com20)
telacom.create_window(1100, 600, window=com24)
root1.mainloop()
